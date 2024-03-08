import os
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import Border, Side, Font, Alignment

from openpyxl import styles
import copy
import pandas as pd
import openpyxl as opx
from openpyxl import styles
import openpyxl.styles as styles
from openpyxl import styles
import copy
import sys
from typing import Dict, List
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
import openpyxl as opx
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time
import shutil
import pandas as pd
import numpy as np
import xlwings as xw
import string
import secrets
import win32com.client as win32
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QScrollArea, QGridLayout
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout
import sys
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtChart import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPainter
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtChart import (QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis)
from src.utils import log_print

try:
    with open('addition/путь.txt', 'r', encoding="utf-8") as file:
        put = file.readline()
    # Путь к первому документу
    file1 = f'{put}/конкурс.xlsx'
    # Путь ко второму документу
    file2 = "addition/конкурс.xlsx"

    # Получаем информацию о дате изменения первого документа
    file1_mtime = os.path.getmtime(file1)
    file1_modified_date = datetime.fromtimestamp(file1_mtime)

    # Получаем информацию о дате изменения второго документа
    file2_mtime = os.path.getmtime(file2)
    file2_modified_date = datetime.fromtimestamp(file2_mtime)

    if file1_modified_date > file2_modified_date:
        src_file = file1
        dst_folder = file2
        shutil.copy(src_file, dst_folder)
        log_print('Copy to local конкурс')
    else:
        src_file = file2
        dst_folder = file1
        shutil.copy(src_file, dst_folder)
        log_print('Copy to X конкурс')
except Exception as e:
    log_print(e, 'Не найдет диск X')
    try:
        with open('addition/путь.txt', 'r', encoding="utf-8") as file:
            file.readline()
            put = file.readline()
        # Путь к первому документу
        file1 = f'{put}/конкурс.xlsx'
        # Путь ко второму документу
        file2 = "addition/конкурс.xlsx"

        # Получаем информацию о дате изменения первого документа
        file1_mtime = os.path.getmtime(file1)
        file1_modified_date = datetime.fromtimestamp(file1_mtime)

        # Получаем информацию о дате изменения второго документа
        file2_mtime = os.path.getmtime(file2)
        file2_modified_date = datetime.fromtimestamp(file2_mtime)

        if file1_modified_date > file2_modified_date:
            src_file = file1
            dst_folder = file2
            shutil.copy(src_file, dst_folder)
            log_print('Copy to local конкурс')
        else:
            src_file = file2
            dst_folder = file1
            shutil.copy(src_file, dst_folder)
            log_print('Copy to X конкурс')
    except:
        log_print(f'{e} Не найдена локальная сеть')


class LoginApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1920, 1080)
        self.setWindowTitle('Проверка логина и пароля')

        self.pushButton = QPushButton("Войти", self)
        self.pushButton.setGeometry(QtCore.QRect(843, 737, 253, 58))
        self.pushButton.setToolTip("<h3>Пройти верификацию</h3>")
        self.pushButton.clicked.connect(self.cheklogpas)
        self.pushButton.setStyleSheet("background-color: rgb(33, 53, 89);\n"
                                      "color: white;\n"
                                      "font: 16pt Myriad pro;\n"
                                      "font-weight: bold;\n"
                                      "\n" "border: 0px solid rgb(6, 73, 129);\n" "border-radius: 15px;")

        self.label2 = QLabel("Введите логин", self)
        self.label2.setGeometry(QtCore.QRect(745, 438, 450, 50))
        self.label2.setAcceptDrops(True)
        self.label2.setAutoFillBackground(False)
        self.label2.setScaledContents(True)
        # self.label2.setAlignment(QtCore.Qt.AlignCenter)
        self.label2.setWordWrap(True)
        self.label2.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
            "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Myriad pro\";\n" "\n" 
                                 "font-weight: bold")

        self.label3 = QLabel("Введите пароль", self)
        self.label3.setGeometry(QtCore.QRect(745, 578, 450, 50))
        self.label3.setAcceptDrops(True)
        self.label3.setAutoFillBackground(False)
        self.label3.setScaledContents(True)
        # self.label3.setAlignment(QtCore.Qt.AlignCenter)
        self.label3.setWordWrap(True)
        self.label3.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
            "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Myriad pro\";\n" "\n" 
                                 "font-weight: bold")
        self.login = '0'
        self.edit_login = QLineEdit(self)
        self.edit_login.setGeometry(QtCore.QRect(745, 485, 442, 55))
        self.edit_login.setObjectName("<h3>Start the Session</h3>")
        self.edit_login.setAlignment(QtCore.Qt.AlignCenter)
        self.edit_login.setStyleSheet(
            "\n" "background-color: rgb(239, 239, 238);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 0px solid rgb(217, 217, 217);\n" "border-radius: 8px;")

        self.text2 = '0'
        self.edit_password = QLineEdit(self)
        self.edit_password.setGeometry(QtCore.QRect(745, 625, 442, 55))
        self.edit_password.setObjectName("<h3>Start the Session</h3>")
        self.edit_password.setAlignment(QtCore.Qt.AlignCenter)
        self.edit_password.setStyleSheet(
            "\n" "background-color: rgb(239, 239, 238);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 0px solid rgb(217, 217, 217);\n" "border-radius: 8px;")

        self.show()

    def cheklogpas(self):
        login = self.edit_login.text()
        password = self.edit_password.text()

        excel_file = 'addition/конкурс.xlsx'

        try:
            df = pd.read_excel(excel_file, sheet_name='Конкурс (Комиссия)')
            match = df[(df.iloc[:, 4] == login) & (df.iloc[:, 5] == password)]
            if not match.empty:
                QMessageBox.information(self, 'Успех', 'Вход выполнен успешно!')
                FIO = match["ФИО"].iloc[0]
                print(FIO)
                self.otkrit(FIO)
            else:
                QMessageBox.warning(self, 'Ошибка', 'Неверный логин или пароль.')

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Ошибка чтения Excel-файла: {str(e)}')

    def otkrit(self, FIO):
        global proverka
        proverka = 0
        self.w = Main_window(FIO)
        app.setStyleSheet(stylesheet976)
        self.w.showMaximized()
        self.w.show()
        self.hide()

class Main_window(QMainWindow):
    def __init__(self, FIO):
        super().__init__()
        self.title = 'Члены конкурса'
        self.FIO = FIO
        self.initUI()

        log_print('Инициализировано главное окно')

    def initUI(self):
        self.setWindowTitle(self.title)

        self.table_widget = QTabWidget(self)
        self.setCentralWidget(self.table_widget)

        self.tab_komic = ExcelTableWidget1(self, 3)
        self.tab_rabot = ExcelTableWidget2(self, 3, self.FIO)

        self.table_widget.addTab(self.tab_komic, "Члены конкурса (комиссия)")
        self.table_widget.addTab(self.tab_rabot, "Члены конкурса (работники)")

        self.table_widget.setStyleSheet('''
            QTabWidget::pane {
                border-top: 3px solid #C2C7CB;
            }
            QTabWidget::tab-bar {
                left: 15px;
            }
            QTabBar::tab {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                            stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
                border: 2px solid #C4C4C3;
                border-bottom-color: #C2C7CB;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 8ex;
                padding: 2px;
            }
            QTabBar::tab:selected, QTabBar::tab:hover {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #fafafa, stop: 0.4 #f4f4f4,
                                            stop: 0.5 #e7e7e7, stop: 1.0 #fafafa);
            }
            QTabBar::tab:selected {
                border-color: #9B9B9B;
                border-bottom-color: #C2C7CB;
            }
            QTabBar::tab:!selected {
                margin-top: 2px;
            }



        ''')
        self.setStyleSheet('''
        QPushButton{
                height: 30px;
                width: 500px;
                margin-top: 10px;
                background-color: rgb(245, 193, 117);
                font: 14pt Times New Roman;
                border: 2px solid rgb(96, 124, 173);
                border-radius: 10px;
                margin: 0;
            }

        QPushButton:hover {
            background-color: rgb(237, 182, 18);
        }      

        QPushButton#blue_button {

            background-color: rgb(182, 202, 237);

        }
        QPushButton#blue_button:hover {

            background-color: rgb(149, 178, 228);

        }

        QHeaderView::section {
            background-color: rgb(240, 240, 240);
        }
        '''
                           )

        self.showMaximized()

class ExcelTableWidget1(QWidget):
    def __init__(self, parent, i):
        super().__init__(parent)
        self.layout1 = QVBoxLayout(self)
        self.table_widget1 = QTableWidget(self)
        self.table_widget1.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget1.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget1.setFont(font)
        self.layout1.addWidget(self.table_widget1)
        self.setLayout(self.layout1)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelik_komic()
        self.show()

    def zapoln_chelik_komic(self):
        df = pd.read_excel('addition/конкурс.xlsx', sheet_name='Конкурс (Комиссия)')
        # Установка количества строк и столбцов в QTableWidget
        df = df[['Статус участника конкурсной комиссии', 'ФИО', 'Должность', 'Подразделение']]

        self.table_widget1.setRowCount(df.shape[0])
        self.table_widget1.setColumnCount(df.shape[1])
        self.table_widget1.setColumnCount(4)

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                xxx = str(df.iat[row, col])
                item = QTableWidgetItem(xxx)
                self.table_widget1.setItem(row, col, item)

        self.table_widget1.resizeColumnsToContents()
        self.table_widget1.setHorizontalHeaderLabels(["Статус участника конкурсной комиссии", "ФИО", "Должность", "Отдел"])
        self.table_widget1.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        self.table_widget1.setColumnWidth(0, 350)
        self.table_widget1.setColumnWidth(1, 300)
        self.table_widget1.setColumnWidth(2, 300)
        self.table_widget1.setColumnWidth(3, 750)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()
        log_print('Таблица конкурса (Комисия) успешно заполнена')


class ExcelTableWidget2(QWidget):
    def __init__(self, parent, i, FIO):
        super().__init__(parent)
        self.layout2 = QVBoxLayout()
        self.layout_h2 = QHBoxLayout()
        self.table_widget2 = QTableWidget(self)
        self.table_widget2.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget2.setFont(font)

        self.FIO = FIO
        self.calendarr = QCalendarWidget()
        self.calendarr.setGridVisible(True)
        self.calendarr.setStyleSheet("QCalendarWidget QToolButton"
                                             "{"
                                             "background-color : lightgrey;"
                                             "color : black"
                                             "}")

        self.calendarchik = QtWidgets.QDateEdit(calendarPopup=True)
        self.calendarchik.setCalendarWidget(self.calendarr)
        self.calendarchik.setDateTime(QtCore.QDateTime.currentDateTime())
        self.calendarchik.setDisplayFormat('dd.MM.yyyy')
        self.calendarchik.setFixedWidth(100)
        self.layout_h2.addWidget(self.calendarchik)

        self.btn = QPushButton(self)
        self.btn.clicked.connect(self.save_to_excel)
        self.btn.setText('Создать бланки комиссии')
        self.layout_h2.addWidget(self.btn)

        self.layout2.addWidget(self.table_widget2)
        self.layout2.addLayout(self.layout_h2)
        self.setLayout(self.layout2)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelick_rabot()
        self.show()

    def zapoln_chelick_rabot(self):

        df = pd.read_excel('addition/конкурс.xlsx', sheet_name='Конкурс (Работники)')
        # Установка количества строк и столбцов в QTableWidget
        df['Процент правильных ответов'] = df['Количество ответов'] / df['Количество вопросов'] * 100
        df = df[['Новый отдел', 'Претендующая Должность', 'ФИО', 'Количество вопросов', 'Количество ответов',\
                 'Процент правильных ответов', self.FIO + " Балл", self.FIO + " ЗаПротив", self.FIO + " Мотивировка"]]
        self.table_widget2.setRowCount(df.shape[0])
        self.table_widget2.setColumnCount(df.shape[1])
        self.table_widget2.setColumnCount(10)

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                if col == 5:
                    xxxi = str(df.iat[row, col])
                    xxxi = int(float(xxxi) + (0.5 if float(xxxi) > 0 else -0.5))
                    item = QTableWidgetItem(xxxi)
                    item.setData(Qt.EditRole, xxxi)
                    self.table_widget2.setItem(row, col, item)
                else:
                    xxx = str(df.iat[row, col])
                    item = QTableWidgetItem(xxx)
                    self.table_widget2.setItem(row, col, item)
        self.table_widget2.resizeColumnsToContents()
        self.table_widget2.setHorizontalHeaderLabels(
            ["Подразделение", "Должность", "ФИО", "Кол-во вопросов", "Кол-во ответов", "% правильных ответов", "Баллы",\
             self.FIO + " Балл", self.FIO + " ЗаПротив", self.FIO + " Мотивировка"])
        self.table_widget2.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
        for row in range(self.table_widget2.rowCount()):
            for col in range(self.table_widget2.columnCount()):
                if col == 6:
                    zzz = self.table_widget2.item(row, 5).text()
                    zzz = int(zzz)
                    zzz = zzz / 10
                    zzz = int(zzz)
                    item = QTableWidgetItem(zzz)
                    item.setData(Qt.EditRole, zzz)
                    self.table_widget2.setItem(row, col, item)
        self.table_widget2.setColumnWidth(0, 750)
        self.table_widget2.setColumnWidth(1, 220)
        self.table_widget2.setColumnWidth(2, 300)
        self.table_widget2.setColumnWidth(3, 150)
        self.table_widget2.setColumnWidth(4, 150)
        self.table_widget2.setColumnWidth(5, 200)
        self.table_widget2.setColumnWidth(6, 60)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()
        log_print('Таблица конкурса (Комисия) успешно заполнена')

    def format_name(self, full_name):
        names = full_name.split()
        surname = names[0]
        names = names[1:3]
        initials = [name[0] + '.' for name in names]
        formatted_name = surname + ' ' + ' '.join(initials)

        return formatted_name

    def smena(self, init):
        names = init.split()
        temp = names[0]
        names = names[1:3]
        names.append(temp)
        formatted_name_2 = ' '.join(names)

        return

    def save_to_excel(self):

        # файлы, которые создаются в этой функции, находятся в files


        # загрузка файла, где находятся все данные
        excel_file = pd.ExcelFile('addition/конкурс.xlsx', engine='openpyxl')

        # извлекаем данные из файла
        df_commission = pd.read_excel(excel_file, sheet_name='Конкурс (Комиссия)')
        df_commission = df_commission[['ФИО']]
        df_workers = excel_file.parse(sheet_name='Конкурс (Работники)')

        # группировка данных по отделу и должности с подсчетом кол-ва записей
        unique_combinations = df_workers.groupby([df_workers.columns[3], df_workers.columns[4]]).size().reset_index(
            name='count')

        for row in range(df_commission.shape[0]):
            # инициализация нового файла эксель на основе шаблона
            wb = opx.load_workbook('addition/St.xlsx')
            ws = wb.active

            #объединение ячеек
            ws.merge_cells(f'A3:C3')
            ws.merge_cells(f'A6:C6')
            ws.merge_cells(f'A7:C7')

            # вставка выбранной даты из календаря в а3
            selected_date = self.calendarchik.date().toString('dd.MM.yyyy')
            ws['A3'] = selected_date

            # извлечение полного имени члена комиссии и форматирование
            full_name = df_commission.iloc[row, 0]
            ws.title = full_name
            ws['A25'] = full_name # это для полного фио в документе (где подпись)

            name = self.format_name(df_commission.iloc[row, 0])
            ws.title = name # это наменования файла эксель

            # извлечение уник. комбинаций
            unique_combinations_set = set()

            # интеграция по уник. комб.
            for index, combination in unique_combinations.iterrows():
                department = combination[df_workers.columns[3]]
                position = combination[df_workers.columns[4]]
                count = combination['count']

                # наименование листа (но там есть типа зно и зно1, это не учитывала , ведь там же еще будет отдел)
                short_position = '.'.join(word[0] for word in department.split())

                # проверка условий для создания нового листа
                if count == 1 and (department, position) not in unique_combinations_set:
                    unique_combinations_set.add((department, position))

                    new_ws = wb.copy_worksheet(ws)
                    new_ws.title = f"{short_position}" #название листа

                    new_ws['A6'] = f"{department}" #вставка должности
                    new_ws['A7'] = f"{position}" # вставка отвела

                elif count > 1:
                    new_ws = wb.copy_worksheet(ws)
                    new_ws.title = f"{short_position}" # название листа

                    new_ws['A6'] = f"{department}" #вставка должности
                    new_ws['A7'] = f"{position}" # вставка отвела

                # извлечение фамилий сотрудников, соответ. отделу и должности
                matching_surnames = df_workers.loc[(df_workers[df_workers.columns[3]] == department) &
                                                   (df_workers[df_workers.columns[4]] == position),
                df_workers.columns[2]].tolist()

                # заполнение фио сотрудников
                new_ws['A23'] = '\n'.join(matching_surnames)

                template_range_below_a23 = new_ws['A25':'A26'] # извлечение ячеек из листа
                template_values_below_a23 = [cell[0].value for cell in template_range_below_a23] # извлечение значений из ячеек (там фио члена комиссии и надпись снизу)
                for cell in template_range_below_a23:
                    cell[0].value = "" # отчистка значений в а25:а26

                # заполнение фио сотрудников начиная с а23 (каждая в новой ячейке)
                start_row = 23
                for idx, surname in enumerate(matching_surnames, start=start_row):
                    new_ws[f'A{idx}'] = surname

                #сдвиг строк ниже данных на кол-во фио сотрудников в таблице и
                #+2 пустые строчки (это применяется ниже в коде!)
                rows_to_shift_down = len(matching_surnames) + 2

                # заполнение фио члена комиссии (учитывая фио сотрудников и +2)
                for idx, value in enumerate(template_values_below_a23):
                    new_ws[f'A{start_row + rows_to_shift_down + idx}'] = value

                # настройка границ для таблицы с фио сотрудников и выравнивание
                less_bold_border = Border(left=Side(border_style='medium', color='000000'),
                                          right=Side(border_style='medium', color='000000'),
                                          top=Side(border_style='medium', color='000000'),
                                          bottom=Side(border_style='medium', color='000000'))

                center_alignment = Alignment(horizontal='center', vertical='center')

                # применение стилей к соответ. ячейкам
                for idx in range(start_row, start_row + len(matching_surnames)):
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        cell = new_ws[f'{col}{idx}']
                        cell.border = less_bold_border
                        cell.alignment = center_alignment
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='Times New Roman', size=12)

            file_name = f'files/{name}.xlsx'
            wb.remove(ws)
            wb.save(file_name)
            wb.close()

stylesheet976 = """
    LoginApp {
        background-image: url(картинки/fon3.png); 
        background-repeat: no-repeat; 
        background-position: center;
    }
"""
            
if __name__ == '__main__':
    log_print('Приложение запущено')
    print(sys.argv)

    if '-t' in sys.argv:
        print('Test open')
        app = QApplication(sys.argv)
        w = Main_window()
        w.show()
        sys.exit(app.exec_())
    else:
        app = QApplication(sys.argv)
        app.setStyleSheet(stylesheet976)
        w = LoginApp()
        w.show()
        # ex = App1()
        sys.exit(app.exec_())
    log_print('Приложение остановлено')