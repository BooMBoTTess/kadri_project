import os
import sys
from typing import Dict, List
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
import random
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from src.documents import create_order
import time
import shutil
import pandas as pd
import numpy as np
import xlwings as xw
import win32com.client as win32
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QScrollArea
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout
import sys
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtChart import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPainter
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtChart import (QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis)

from src.utils import log_print
import src.Accept_order as Accept_order
from src.konkurs import KonkursTabulation

"""Приложение для кадров"""


try:
    with open('addition/путь.txt', 'r', encoding="utf-8") as file:
        put = file.readline().strip('\n')
    # Путь к первому документу
    file1 = f'{put}/kadrifile.xlsx'
    # Путь ко второму документу
    file2 = "addition/kadrifile.xlsx"

    # Получаем информацию о дате изменения первого документа
    file1_mtime = os.path.getmtime(file1)
    file1_modified_date = datetime.fromtimestamp(file1_mtime)

    # Получаем информацию о дате изменения второго документа
    file2_mtime = os.path.getmtime(file2)
    file2_modified_date = datetime.fromtimestamp(file2_mtime)

    if file1_modified_date > file2_modified_date:
        src_file = file1
        dst_folder = file2
        shutil.copy(src_file, dst_folder)
        log_print('Copy to local kadrifile')
    else:
        src_file = file2
        dst_folder = file1
        shutil.copy(src_file, dst_folder)
        log_print('Copy to X kadrifile')
except Exception as e:
    log_print(e, 'Не найдет диск X')
    try:
        with open('addition/путь.txt', 'r', encoding="utf-8") as file:
            file.readline()
            put = file.readline()
        # Путь к первому документу
        file1 = f'{put}/kadrifile.xlsx'
        # Путь ко второму документу
        file2 = "addition/kadrifile.xlsx"

        # Получаем информацию о дате изменения первого документа
        file1_mtime = os.path.getmtime(file1)
        file1_modified_date = datetime.fromtimestamp(file1_mtime)

        # Получаем информацию о дате изменения второго документа
        file2_mtime = os.path.getmtime(file2)
        file2_modified_date = datetime.fromtimestamp(file2_mtime)

        if file1_modified_date > file2_modified_date:
            src_file = file1
            dst_folder = file2
            shutil.copy(src_file, dst_folder)
            log_print('Copy to local kadrifile')
        else:
            src_file = file2
            dst_folder = file1
            shutil.copy(src_file, dst_folder)
            log_print('Copy to X kadrifile')
    except:
        log_print(f'{e} Не найдена локальная сеть')

app = xw.App(visible=False)
wb = xw.Book('addition/kadrifile.xlsx')
wb.save('addition/kadrifile.xlsx')
wb.close()
app.quit()


class Login_page_app(QMainWindow):
    def __init__(self):
        super().__init__()
        self.title = 'Авторизация отдел кадров'

        self.pushButton = QPushButton("ВОЙТИ", self)
        self.pushButton.setGeometry(QtCore.QRect(843, 737, 253, 58))
        self.pushButton.setToolTip("<h3>Пройти верификацию</h3>")
        self.pushButton.clicked.connect(self.cheklogpas)
        self.pushButton.setStyleSheet("background-color: rgb(33, 53, 89);\n"
                                      "color: white;\n"
                                      "font: 16pt Myriad pro;\n"
                                      "font-weight: bold;\n"
                                      "\n" "border: 0px solid rgb(6, 73, 129);\n" "border-radius: 15px;")

        self.label2 = QLabel("Введите логин", self)
        self.label2.setGeometry(QtCore.QRect(745, 438, 450, 50))
        self.label2.setAcceptDrops(True)
        self.label2.setAutoFillBackground(False)
        self.label2.setScaledContents(True)
        # self.label2.setAlignment(QtCore.Qt.AlignCenter)
        self.label2.setWordWrap(True)
        self.label2.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Myriad pro\";\n" "\n"
                                  "font-weight: bold")

        self.label3 = QLabel("Введите пароль", self)
        self.label3.setGeometry(QtCore.QRect(745, 578, 450, 50))
        self.label3.setAcceptDrops(True)
        self.label3.setAutoFillBackground(False)
        self.label3.setScaledContents(True)
        # self.label3.setAlignment(QtCore.Qt.AlignCenter)
        self.label3.setWordWrap(True)
        self.label3.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Myriad pro\";\n" "\n"
                                  "font-weight: bold")

        self.text1 = '0'
        self.lineEdit = QLineEdit(self)
        self.lineEdit.setGeometry(QtCore.QRect(745, 485, 442, 55))
        self.lineEdit.setObjectName("<h3>Start the Session</h3>")
        self.lineEdit.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit.setStyleSheet(
            "\n" "background-color: rgb(239, 239, 238);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 0px solid rgb(217, 217, 217);\n" "border-radius: 8px;")

        self.text2 = '0'
        self.lineEdit2 = QLineEdit(self)
        self.lineEdit2.setGeometry(QtCore.QRect(745, 625, 442, 55))
        self.lineEdit2.setObjectName("<h3>Start the Session</h3>")
        self.lineEdit2.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit2.setStyleSheet(
            "\n" "background-color: rgb(239, 239, 238);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 0px solid rgb(217, 217, 217);\n" "border-radius: 8px;")

        self.text3 = '0'
        self.lineEdit3 = QLineEdit(self)
        self.lineEdit3.setGeometry(QtCore.QRect(580, 280, 730, 50))
        self.lineEdit3.setObjectName("<h3>Start the Session</h3>")
        self.lineEdit3.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit3.setStyleSheet(
            "\n" "background-color: rgb(217, 217, 217);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 0px solid rgb(217, 217, 217);\n" "border-radius: 10px;")
        self.lineEdit3.setVisible(False)

        self.pushButton2 = QPushButton("", self)
        self.pushButton2.setGeometry(QtCore.QRect(1385, 800, 80, 80))
        self.pushButton2.clicked.connect(self.opendoor)
        self.pushButton2.setStyleSheet(
            "\n" "background-color: rgba(0, 0, 0, 0);\n" "\n" "font: 10pt \"Times New Roman\";"
            "\n" "border-radius: 12px;")

        self.pushButton7 = QPushButton("Поменять путь выгрузки файлов", self)
        self.pushButton7.setGeometry(QtCore.QRect(580, 200, 700, 50))
        self.pushButton7.clicked.connect(self.change_opendoor)
        self.pushButton7.setStyleSheet(
            "\n" "background-color: rgb(127, 155, 205);\n" "\n" "font: 18pt \"Times New Roman\";"
            "\n" "border: 3px solid rgb(6, 73, 129);\n" "border-radius: 30px;")
        self.pushButton7.setVisible(False)

        self.showMaximized()
        self.show()

    def opendoor(self):
        self.text1 = self.lineEdit.text()
        self.text2 = self.lineEdit2.text()
        self.loginn1 = 'admin'
        self.loginn2 = 'admin'
        if self.loginn1 == self.text1 and self.loginn2 == self.text2:
            self.lineEdit3.setVisible(True)
            self.pushButton7.setVisible(True)

    def change_opendoor(self):
        self.putfiles = self.lineEdit3.text()
        try:
            with open('file.txt', 'w') as file:
                file.truncate(0)
            with open("addition/путь.txt", "w", encoding="utf-8") as file:
                file.write(self.putfiles)

        except FileNotFoundError:
            log_print('Файл логин пароль не найден. ', e)
            QMessageBox.warning(self, "Ошибка777", "Файл проверки лог/пас не найден.", '9')

    def cheklogpas(self):
        global glav
        glav = 0
        self.text1 = self.lineEdit.text()
        self.text2 = self.lineEdit2.text()
        self.r = 0
        self.loginn1 = 'admin'
        self.loginn2 = 'admin'
        try:
            with open("addition/логин.txt", "r", encoding="utf-8") as file:
                for line in file:
                    # Отделяем первый набор символов до знака "№"
                    login = line.split("/")[0]
                    password = line.split("/")[1]
                    global otdel
                    otdel = line.split("/")[2]
                    # Сравниваем с поисковым запросом
                    if login == self.text1 and password == self.text2:
                        # Если найдено соответствие, выводим строчку
                        QMessageBox.information(self, "Найдено соответствие", line)
                        log_print(f'user: {login, password} зашел. ')
                        if self.loginn1 == self.text1 and self.loginn2 == self.text2:
                            glav = 1
                        self.otkrit()
                        self.r = 1
                        break

                if self.r == 0:
                    QMessageBox.warning(self, "Ошибка 333", 'Ошибка вы ввели неверный логин или пароль')
                else:
                    pass
        except FileNotFoundError as e:
            log_print('Логпас не найден', {e})
            QMessageBox.warning(self, f' Ошибка777", "Файл проверки лог/пас не найден.')

    def otkrit(self):
        global proverka
        proverka = 0
        self.w = Main_window()
        app.setStyleSheet(stylesheet976)
        self.w.showMaximized()
        self.w.show()
        self.hide()


class Main_window(QMainWindow):
    def __init__(self):
        super().__init__()
        self.title = 'Отдел кадров'
        self.initUI()

        self.func_onChange = {0: self.tab_form_order.zapoln_chelik,
                              4: self.tab_accept_order.update_orders
                              }

        log_print('Инициализировано главное окно')

    def initUI(self):
        self.setWindowTitle(self.title)

        self.table_widget = QTabWidget(self)
        self.setCentralWidget(self.table_widget)
        self.df = pd.read_excel('addition/kadrifile.xlsx')

        self.tab_form_order = ExcelTableWidget1(self.df, self, 3)
        self.tab_list_tools = ExcelTableWidget(self.df, self, 3)
        self.tab_list_staff = ExcelTableWidget2(self.df, self, 3)
        self.tab_email_throw = ExcelTableWidget3(self.df, self, 3)
        self.tab_tool5 = ExcelTableWidget5(self.df, self, 4)
        self.tab_help = ExcelTableWidget4(self.df, self, 3)
        self.tab_conc = KonkursTabulation(self.df)
        self.tab_accept_order = Accept_order.Accept_order_tab(self)

        self.table_widget.addTab(self.tab_form_order, "Формирование приказа")
        self.table_widget.addTab(self.tab_list_tools, "Лист инструментов")
        self.table_widget.addTab(self.tab_list_staff, "Общий лист штатки")
        self.table_widget.addTab(self.tab_email_throw, "Лист рассылки")
        self.table_widget.addTab(self.tab_accept_order, "Приказы в работе")
        self.table_widget.addTab(self.tab_conc, "Конкурс")
        self.table_widget.addTab(self.tab_tool5, "Дашборд")
        self.table_widget.addTab(self.tab_help, "Помощь")

        self.table_widget.currentChanged.connect(self.onChange)
        self.table_widget.setStyleSheet('''
            QTabWidget::pane {
                border-top: 3px solid #C2C7CB;
            }
            QTabWidget::tab-bar {
                left: 15px;
            }
            QTabBar::tab {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                            stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
                border: 2px solid #C4C4C3;
                border-bottom-color: #C2C7CB;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 8ex;
                padding: 2px;
            }
            QTabBar::tab:selected, QTabBar::tab:hover {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #fafafa, stop: 0.4 #f4f4f4,
                                            stop: 0.5 #e7e7e7, stop: 1.0 #fafafa);
            }
            QTabBar::tab:selected {
                border-color: #9B9B9B;
                border-bottom-color: #C2C7CB;
            }
            QTabBar::tab:!selected {
                margin-top: 2px;
            }
            
                                
                        
        ''')
        self.setStyleSheet('''
        QPushButton{
                height: 30px;
                width: 500px;
                margin-top: 10px;
                background-color: rgb(245, 193, 117);
                font: 14pt Times New Roman;
                border: 2px solid rgb(96, 124, 173);
                border-radius: 10px;
                margin: 0;
            }
                        
        QPushButton:hover {
            background-color: rgb(237, 182, 18);
        }      
        
        QPushButton#blue_button {
           
            background-color: rgb(182, 202, 237);
            
        }
        QPushButton#blue_button:hover {
           
            background-color: rgb(149, 178, 228);
            
        }
        
        QHeaderView::section {
            background-color: rgb(240, 240, 240);
        }
        '''
                           )

        self.showMaximized()

    def onChange(self, tab_page: int):
        """
        Функция, которая срабатывает при переходе на другую строку, вызывает функцию из словаря func_onchange.
        Для того, чтобы добавить новую табу. нужно добавить к этому словарю номер этой табы и функцию,
        которая должна отрабатываться при переходе на нее
        :param tab_page: int: число вкладки куда переключаемся.
        :return:
        """
        if tab_page in self.func_onChange.keys():
            self.func_onChange[tab_page]()


class ExcelTableWidget1(QWidget):
    """
    Формирование приказа. Первое окно, формирующее приказы.
    При нажатии формирования приказа запускает def izmenit_cheliki
    Потом отккрывается класс Information HR_sheet. при нажатии на кнопку -> def start_work_order.
    Порядок работы В try\except:
        1. Запускаем прогресс бар
        2. Забираем данные с нижней таблицы def get_values_from_BottomSheet
        2. Чекаем правильно ли заполнены данные.
            2.1 Если они не заполнены правильно error raise, который принимается в except здесь.
            Ошибка идет в логи и в окно, все останавливается
        3. Отправляем приказы в генерацию документа ворд
        4. Отправляем приказы в лист ['Приказы в работе']
        5. fill_potential_close - закрываем единичками в первом столбике должности, на которые претендуют
        6. zapoln_chelik - ХЗ что делает.


    """

    def __init__(self, df, parent, i):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.h_layout2 = QHBoxLayout()
        self.h_layout = QHBoxLayout()
        self.df = df
        self.table_widget23 = QTableWidget(self)
        self.table_widget23.setEditTriggers(QTableWidget.AllEditTriggers)

        self.table_widget9999 = QTableWidget(self)
        self.table_widget9999.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget23.setStyleSheet("background-image: url(картинки/2.png)")
        self.table_widget9999.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget23.setFont(font)
        self.table_widget9999.setFont(font)

        self.label1 = QLabel("Поиск:")
        self.label1.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                  "font-weight: bold")

        self.label33 = QLabel("Перечень сотрудников добавленных в приказ:")
        self.label33.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                   "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                   "font-weight: bold")

        self.searchLineEdit = QLineEdit()
        self.searchLineEdit.textChanged.connect(self.search_table)

        self.button_addcheliki = QPushButton("Добавить сотрудника в приказ")
        self.button_addcheliki.clicked.connect(self.add_cheliki)
        self.button_addcheliki.setObjectName('blue_button')

        self.button_deletecheliki = QPushButton("Удалить сотрудника из приказа")
        self.button_deletecheliki.clicked.connect(self.delete_cheliki)
        self.button_deletecheliki.setObjectName('blue_button')

        self.button_izmencheliki = QPushButton("Начать работу над приказом")
        self.button_izmencheliki.clicked.connect(self.izmenit_cheliki)

        self.h_layout.addWidget(self.button_addcheliki)
        self.h_layout.addWidget(self.button_deletecheliki)

        self.h_layout2.addWidget(self.label1)
        self.h_layout2.addWidget(self.searchLineEdit)

        self.layout.addLayout(self.h_layout2)
        self.layout.addLayout(self.h_layout)

        self.layout.addWidget(self.button_izmencheliki)
        self.layout.addWidget(self.table_widget23)
        self.layout.addWidget(self.label33)
        self.layout.addWidget(self.table_widget9999)

        self.setLayout(self.layout)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelik()
        log_print('Инициализирована 1 вкладка')

        self.show()

    def check_cheliki(self):
        try:
            dfcheck = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')

            columnnames = ['Unnamed: 0', 'Подразделение', 'Должность', 'ФИО', 'Статус', 'Новый отдел', 'Дата',
                           'Номер приказа']
            dataframe006 = pd.DataFrame(columns=columnnames)
            for row in range(self.table_widget9999.rowCount()):
                rowData = [
                    self.table_widget9999.item(row, col).text() if col != 3 else self.table_widget9999.cellWidget(row,
                                                                                                                  col).currentText()
                    for col
                    in range(self.table_widget9999.columnCount())]
                rowData.insert(0, " ")
                dataframe006 = pd.concat([dataframe006, pd.DataFrame([rowData], columns=columnnames)])
                row_color = QColor(0, 255, 0)
                self.table_widget9999.item(row, 2).setBackground(row_color)

            for index, row in dataframe006.iterrows():
                match = dfcheck[(dfcheck['Подразделение'] == row['Подразделение']) & (
                        dfcheck['Должность'] == row['Должность']) & (dfcheck['ФИО'] == row['ФИО'])]

                if not match.empty:
                    dfcheck.loc[match.index, 'ФИО'] = np.nan

                match2 = dfcheck[
                    (dfcheck['Подразделение'] == row['Новый отдел']) & (dfcheck['Должность'] == row['Статус']) & (
                        dfcheck['ФИО'].isnull())]

                if not match2.empty:
                    for k in range(len(match2.index)):
                        if row['Статус'] != '':
                            dfcheck.loc[match2.index[k], 'ФИО'] = row['ФИО']
                            dfcheck.loc[match2.index[k], 'Дата'] = row['Дата']
                            dfcheck.loc[match2.index[k], 'Номер приказа'] = row['Номер приказа']

                            break

            for index, row in dataframe006.iterrows():
                content = row['Подразделение']
                content_2 = row['Статус']
                first_row = dfcheck[(dfcheck['Подразделение'] == content) & (dfcheck['Должность'] == content_2)].index[
                    0]
                last_row = dfcheck[(dfcheck['Подразделение'] == content) & (dfcheck['Должность'] == content_2)].index[
                    -1]
                df_copy = dfcheck[first_row:last_row + 1].copy()
                df_copy.sort_values(by='ФИО', inplace=True)
                dataframe_up = pd.concat([dfcheck.iloc[:first_row], df_copy, dfcheck.iloc[last_row + 1:]], axis=0)
        except:
            self.message_box9 = QMessageBox(self)
            self.message_box9.setWindowTitle("Отчет об отработке")
            self.message_box9.setText("Имеется ошибка в введенных вами данных, исправьте и проверьте заново")
            self.message_box9.setStandardButtons(QMessageBox.Ok)
            self.message_box9.show()

    def delete_cheliki(self):
        row = self.table_widget9999.currentRow()
        if row > -1:  # Если есть выделенная строка/элемент
            self.table_widget9999.removeRow(row)
            # Следующий вызов нужен для того, чтобы
            # сбросить индекс выбранной строки (чтобы currentRow установился в -
            log_print(f'строка {row} удалена из нижнего виджета')
            self.table_widget9999.selectionModel().clearCurrentIndex()

        else:
            current_row_count = self.table_widget9999.rowCount()
            self.table_widget9999.setRowCount(current_row_count - 1)
            log_print(f'строка{row} удалена из нижнего виджета')

    class Information_HR_sheet(QWidget):
        """
            window with information about HR where you can enter name and post by
            HR, Tutor, Law, Admin worker.
            It used in future in documents.py to create order

        """

        def __init__(self, parent):
            super().__init__()
            self.setWindowTitle('Информация о кадрах')
            self.setMinimumWidth(600)
            self.setMinimumHeight(400)


            self.grid_layout = QGridLayout()
            self.setLayout(self.grid_layout)
            self.edited_lines_list = []
            self.edited_combo_box = []
            self.df_staff = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
            self.setStyleSheet("""            
                        QGridLayout{
                            
                        }
                        
                        QLabel{
                            margin-left:20px;
                            color: #383838;
                            font: 16pt Myriad pro;
                            
                        }
                        
                        QLabel#Title_label{
                            font-weight: bold;
                            color: black;
                            margin-bottom:10px;
                            
                        }
                        QLineEdit{
                            margin-right:20px;
                            background-color: white;
                        }
                        
                        QComboBox{
                        background-color: #D3D3D3;
                        }
                        
                        QPushButton{
                            height: 30px;
                            margin-top: 50px;
                            background-color: rgb(245, 193, 117);
                            font: 16pt Myriad pro;
                            border: 2px solid rgb(96, 124, 173);
                            border-radius: 10px;
                        }
                        
                        QTableWidget{
                            background
                        }
                        
                        
                        """)

            self.create_table()
            self.create_button(parent)

        def create_button(self, parent):
            """
            Function that create accept button on information about HR
            :param parent: Папочка
            :return:
            """
            self.accept_button = QPushButton(self)
            self.accept_button.setText('Подтвердить')
            self.accept_button.clicked.connect(parent.start_work_order)

            self.grid_layout.addWidget(self.accept_button, 10, 0, 1, 4)

            self.label12345 = QLabel("Обратите внимание, на заполнение всех полей, и полный формат ФИО", self)
            self.label12345.setAcceptDrops(True)
            self.label12345.setAlignment(Qt.AlignCenter)
            self.label12345.setStyleSheet("\n" "color: rgb(100, 100, 100);\n"
                                          "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 12pt \"Myriad pro\";\n" "\n"
                                          "font-weight: bold")
            self.grid_layout.addWidget(self.label12345, 11, 0, 1, 4)

        def create_table(self):
            """
            Fill grid_table on layout by labels and editlabels
            Cycle that create blocks like:
            Title
            FIO ______ post ________
            """

            titles = ['Сотрудник отдела кадров', 'Куратор отдела кадров \n(Лист согласования)',
                      'Сотрудник юридического отдела\n(Лист согласования)',
                      'Сотрудник административного отдела\n(Лист согласования)',
                      'Руководитель\n(ИО руководителя)']
            list_of_names = self.get_names_for_fill()
            self.lines = []
            for block in range(0, 5):  # цикл, чтобы сделать 4 блока
                post = QComboBox()
                if block != 4:
                    post.addItems(
                        ["", "Должность 1", "Должность 2", "Должность 3"])
                else:
                    post.addItems(
                        ["", "Руководитель", 'И.О. руководителя'])

                name_combobox = QComboBox()
                name_combobox.setEditable(True)
                # combo_2.setWordWrap(True)
                name_combobox.setInsertPolicy(QComboBox().InsertAfterCurrent)
                name_combobox.addItems(list_of_names[block])

                self.lines.append([name_combobox, post])  # Добавляем заполняемые поля


                # добавление виджетов на сетку
                title_label = QLabel(titles[block], self)
                title_label.setObjectName('Title_label')
                title_label.setAlignment(Qt.AlignCenter)

                self.grid_layout.addWidget(title_label, block * 2, 0, 1, 4,
                                           alignment=Qt.AlignmentFlag.AlignBottom)

                self.grid_layout.addWidget(QLabel('ФИО', self), block * 2 + 1, 0)
                self.grid_layout.addWidget(self.lines[block][0], block * 2 + 1, 1)

                self.grid_layout.addWidget(QLabel('Должность', self), block * 2 + 1, 2)
                self.grid_layout.addWidget(self.lines[block][1], block * 2 + 1, 3)

        def get_HR_info_values(self):
            """
            Get values, that filled in HR window
            :return: dict format {'HR_name': '', 'HR_post': '',
                              'Curator_name': '', 'Curator_post': '',
                              'Law_dep_name': '', 'Law_dep_post': '',
                              'admin_dep_name': '', 'admin_dep_post': '',
                                'head_dep_name': '', 'head_dep_post': ''
                              }
            """
            HR_info_values = {'HR_name': '', 'HR_post': '',
                              'Curator_name': '', 'Curator_post': '',
                              'Law_dep_name': '', 'Law_dep_post': '',
                              'admin_dep_name': '', 'admin_dep_post': '',
                              'head_dep_name': '', 'head_dep_post': ''
                              }
            k_lines, k_boxes = 0, 0
            for key in HR_info_values:
                if k_lines == k_boxes:
                    HR_info_values[key] = self.lines[k_lines][0].currentText()
                    k_lines += 1
                else:
                    HR_info_values[key] = self.lines[k_boxes][1].currentText()
                    k_boxes += 1

            return HR_info_values

        def get_names_for_fill(self) -> List[List[str]]:
            HRs = 'Отдел государственной гражданской службы и кадров'
            HR_kurator = ('Начальник отдела', 'Заместитель начальника отдела')
            law_dep = 'Юридический отдел'
            admin = 'Административный отдел'
            heads = 'Руководство'
            result = []
            tmp = self.df_staff[self.df_staff['Подразделение'] == HRs]
            result.append(list(tmp['ФИО'].dropna()))

            tmp = tmp[(tmp['Должность'] == 'Начальник отдела') | (tmp['Должность'] == 'Заместитель начальника отдела')]
            result.append(list(tmp['ФИО'].dropna()))

            tmp = self.df_staff[self.df_staff['Подразделение'] == law_dep]
            result.append(list(tmp['ФИО'].dropna()))

            tmp = self.df_staff[self.df_staff['Подразделение'] == admin]
            result.append(list(tmp['ФИО'].dropna()))

            tmp = self.df_staff[self.df_staff['Подразделение'] == heads]
            result.append(list(tmp['ФИО'].dropna()))

            return result


    def izmenit_cheliki(self):

        try:
            for row in range(self.table_widget9999.rowCount()):
                rowData333 = []
                for col in range(self.table_widget9999.columnCount()):
                    if col != 3 and col != 7 and col != 4 and col != 5 and col != 8:
                        zzz = self.table_widget9999.item(row, col).text()
                        rowData333.append(zzz)
                    elif col == 8 or col == 5:
                        zzz3 = self.table_widget9999.cellWidget(row, col).text()
                        rowData333.append(zzz3)
                    else:
                        zzz2 = self.table_widget9999.cellWidget(row, col).currentText()
                        rowData333.append(zzz2)
                print(rowData333)
            log_print('Открыто окно с информацией об HR')
            self.inf = self.Information_HR_sheet(self)
            self.inf.show()
        except:
            self.message_box9616 = QMessageBox()
            self.message_box9616.setWindowTitle("Отчет об ошибках")
            self.message_box9616.setText(
                "Заполните все ПУСТЫЕ поля согласно предоставленной инструкции!")
            self.message_box9616.setStandardButtons(QMessageBox.Ok)
            self.message_box9616.show()

    def order_to_work(self, bottom_sheet: pd.DataFrame):
        """
        Добавляет приказы в табличук приказы в работе
        :param bottom_sheet:  Нижняя табличка df006
        :return:
        """
        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Приказы в работе']

        fastdf = bottom_sheet.iloc[:, 1:]
        for r in dataframe_to_rows(fastdf, index=False, header=False):
            ws.append(r)
        book.save(f"логи/File XYZ_{datetime.now().strftime('%Y-%m-%d %H.%M.%S')}.xlsx")
        book.save('addition/kadrifile.xlsx')
        book.close()

    def order_to_docs(self, dataframe_bottom: pd.DataFrame, HR_info: Dict):
        """Формирует документы с приказами"""
        word = create_order()
        for index, row in dataframe_bottom.iterrows():
            otdel = row['Подразделение']
            dolzn = row['Должность']
            full_name = row['ФИО']
            new_dolzn = row['Статус']
            new_otdel = row['Новый отдел']
            data_prikaz = row['Дата']
            data_zayav = row['Дата заявления']
            document_type = row['Тип документа']

            word.create_word_file(full_name, dolzn, new_dolzn, otdel, new_otdel,
                                  data_prikaz, data_zayav, HR_info, document_type)
            time.sleep(1)  # Важная штука

    def is_in_staff(self, DataFrame_staff: pd.DataFrame, Dataframe_bottom: pd.DataFrame):
        """Существует ли такой человек"""
        for index, row in Dataframe_bottom.iterrows():
            if (row['Статус'] == 'Уволен') \
                    or (row['Подразделение'] != '') \
                    or (row['Должность'] != ''):
                match = DataFrame_staff[(DataFrame_staff['Подразделение'] == row['Подразделение'])
                                        & (DataFrame_staff['Должность'] == row['Должность'])
                                        & (DataFrame_staff['ФИО'] == row['ФИО'])]
                if match.empty:
                    raise ValueError(f'{row["ФИО"]} - такого человека не существует. ')

    def can_move(self, DataFrame_staff: pd.DataFrame, Dataframe_bottom: pd.DataFrame):
        """Можно ли переместиться на это место. Проверяет и ФИО и Заполненность Unnamed"""
        df_stat_grouped = DataFrame_staff.loc[:, ['Подразделение', 'Должность', 'ФИО', 'Unnamed: 0']]
        df_stat_grouped = df_stat_grouped[(df_stat_grouped['ФИО'].isnull())
                                          & (df_stat_grouped['Unnamed: 0'] != 1)].loc[:,
                          ['Подразделение', 'Должность']] \
            .value_counts(dropna=False).reset_index(name='count')

        df_widget_grouped = Dataframe_bottom.loc[:, ['Новый отдел', 'Статус', 'ФИО']]
        df_widget_grouped = df_widget_grouped.loc[:, ['Новый отдел', 'Статус']] \
            .value_counts(dropna=False).reset_index(name='count').rename(columns={"Новый отдел": "Подразделение",
                                                                                  "Статус": "Должность"})
        for index, row in df_widget_grouped.iterrows():
            if row['Должность'] != 'Уволен':
                free_positions = df_stat_grouped[(df_stat_grouped['Подразделение'] == row['Подразделение'])
                                                 & (df_stat_grouped['Должность'] == row['Должность'])].loc[:, 'count']
                if free_positions.empty:
                    raise ValueError(f'Невозможно переместить данное количество людей. '
                                     f'Не существует свободной должности в {row["Подразделение"], row["Должность"]}. '
                                     f'проверьте правильность написания подразделения и должности.')

                elif int(free_positions.iloc[0]) < int(row['count']):
                    raise ValueError(f'Невозможно переместить данное количество людей. '
                                     f'Количество свободных мест в {row["Подразделение"], row["Должность"]} '
                                     f'меньше, чем задано: {int(free_positions.iloc[0])} < {int(row["count"])}. ')

    def is_order_in_work(self, DataFrame_order: pd.DataFrame, Dataframe_bottom: pd.DataFrame):
        """Существует ли приказ на этого человека"""
        bad_names = []
        if not DataFrame_order.empty:
            for name in Dataframe_bottom['ФИО']:
                match4 = DataFrame_order[DataFrame_order['ФИО'] == name]
                if not match4.empty:
                    bad_names.append(name)
        if bad_names != []:
            raise ValueError(f'Приказ, на данных сотрудников уже существует.\n'
                             f'{bad_names}')

    def Checker_every(self, DataFrame_staff: pd.DataFrame, Dataframe_bottom: pd.DataFrame,
                      DataFrame_order: pd.DataFrame):

        self.is_in_staff(DataFrame_staff, Dataframe_bottom)
        self.can_move(DataFrame_staff, Dataframe_bottom)
        self.is_order_in_work(DataFrame_order, Dataframe_bottom)
        log_print('Check_succesfull')

    def get_values_from_BottomSheet(self):
        '''Подгрузка из нижней таблицы данных'''

        columnnames = ['Unnamed: 0', 'Подразделение', 'Должность', 'ФИО', 'Статус', 'Новый отдел', 'Дата',
                       'Номер приказа', 'Тип документа', 'Дата заявления']
        df_bottom_sheet = pd.DataFrame(columns=columnnames)
        for row in range(self.table_widget9999.rowCount()):
            rowData = []
            for col in range(self.table_widget9999.columnCount()):
                if col != 3 and col != 7 and col != 4 and col != 5 and col != 8:
                    zzz = self.table_widget9999.item(row, col).text()
                    rowData.append(zzz)
                elif col == 8 or col == 5:
                    zzz3 = self.table_widget9999.cellWidget(row, col).text()
                    rowData.append(zzz3)
                else:
                    zzz2 = self.table_widget9999.cellWidget(row, col).currentText()
                    rowData.append(zzz2)

            rowData.insert(0, " ")
            df_bottom_sheet = pd.concat([df_bottom_sheet, pd.DataFrame([rowData], columns=columnnames)])

        return df_bottom_sheet

    def fill_potential_close(self, DataFrame_staff, Dataframe_bottom):
        for index, row in Dataframe_bottom.iterrows():
            match = DataFrame_staff[(DataFrame_staff['Подразделение'] == row['Новый отдел'])
                                    & (DataFrame_staff['Должность'] == row['Статус'])
                                    & (DataFrame_staff['Unnamed: 0'] != 1)
                                    & (DataFrame_staff['ФИО'].isnull())
                                    ]
            for k in range(len(match.index)):
                DataFrame_staff.loc[match.index[k], 'Unnamed: 0'] = 1
                break

        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Штатка']
        ws.delete_cols(1, 8)
        ws.delete_rows(1, 500)
        for r in dataframe_to_rows(DataFrame_staff, index=False, header=True):
            ws.append(r)

        book.save('addition/kadrifile.xlsx')
        book.close()

    def start_work_order(self):
        """
        Создает приказ и документ
        :return:
        """

        log_print('Начата работа над приказом. ')
        HR_info = self.inf.get_HR_info_values()  # забрать HR_info
        self.inf.close()

        df_staff = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:E')
        df_order = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Приказы в работе', usecols='A:H')

        self.progress_dialog = QProgressDialog("Выполнение задач", None, 0, 100)
        self.progress_dialog.setWindowTitle("Ход работы приложения")
        self.progress_dialog.setWindowModality(2)
        self.progress_dialog.setMinimumDuration(100)
        self.progress_dialog.setAutoClose(True)
        self.progress_dialog.setAutoReset(True)

        try:
            df_bottom_sheet = self.get_values_from_BottomSheet()

            self.progress_dialog.setLabelText(f"Выполнение задач:\n"
                                              f"Изучение внесенных данных---| \n")
            self.progress_dialog.setValue(20)
            self.Checker_every(df_staff, df_bottom_sheet, df_order)

            self.progress_dialog.setLabelText(f"Выполнение задач:\n"
                                              f"Изучение внесенных данных---| Выполнено\n"
                                              f"Формирование приказов---|    \n")
            self.progress_dialog.setValue(45)
            self.order_to_docs(df_bottom_sheet, HR_info)

            self.progress_dialog.setLabelText(f"Выполнение задач:\n"
                                              f"Изучение внесенных данных---| Выполнено\n"
                                              f"Формирование приказов---| Выполнено  \n"
                                              f"Приказов в работу---| ")
            self.progress_dialog.setValue(75)
            self.order_to_work(df_bottom_sheet)

            self.progress_dialog.setLabelText(f"Выполнение задач:\n"
                                              f"Изучение внесенных данных---| Выполнено\n"
                                              f"Формирование приказов---| Выполнено  \n"
                                              f"Отправка приказов в работу---| Выполнено \n"
                                              f"Изменение в таблице---| \n"
                                              )
            self.progress_dialog.setValue(80)
            self.fill_potential_close(df_staff, df_bottom_sheet)

            self.progress_dialog.setLabelText(f"Выполнение задач:\n"
                                              f"Изучение внесенных данных---| Выполнено\n"
                                              f"Формирование приказов---| Выполнено  \n"
                                              f"Отправка приказов в работу---| Выполнено \n"
                                              f"Изменение в таблице---| Выполнено \n"
                                              )
            self.progress_dialog.setValue(100)
            self.zapoln_chelik()
        except Exception as e:
            self.progress_dialog.close()
            msg = QMessageBox(self)
            msg.setWindowTitle("Ошибка")
            msg.setText(f"{e}. Невозможно выполнить запрос.")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.show()
            log_print(e, '| Невозможно выполнить запрос.')
            print(e, '| Невозможно выполнить запрос.')

        # Открыли эксельку пересчитала, закрываем
        app = xw.App(visible=False)
        wb = xw.Book('addition/kadrifile.xlsx')
        wb.save('addition/kadrifile.xlsx')
        wb.close()
        app.quit()

        log_print(f'Работа над приказом закончилась. Успешно обработано {len(df_bottom_sheet)} запросов.')

    def add_cheliki(self):
        try:
            current_row_count = self.table_widget9999.rowCount()
            self.table_widget9999.setColumnCount(9)
            self.table_widget9999.setRowCount(current_row_count + 1)
            combo = QComboBox()
            combo.addItems(
                ["", "Должность 1", "Должность 2", "Должность 3"])
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)

            combo_1 = QComboBox()
            combo_1.addItems(["", "Назначение", "Увольнение", "Перевод"])
            self.table_widget9999.setCellWidget(current_row_count, 7, combo_1)

            combo_2 = QComboBox()
            combo_2.setEditable(True)
            # combo_2.setWordWrap(True)
            combo_2.setInsertPolicy(QComboBox().InsertAfterCurrent)
            combo_2.addItems(['', 'Отдел 1','Отдел 2','Отдел 3'Отдел 4',])
            self.table_widget9999.setCellWidget(current_row_count, 4, combo_2)

            selected_row = self.table_widget23.currentRow()
            row_items = []

            for column in range(self.table_widget23.columnCount()):
                item = self.table_widget23.item(selected_row, column)
                row_items.append(item.text())
            row_position = self.table_widget9999.rowCount()

            for column, item in enumerate(row_items):
                self.table_widget9999.setItem(row_position - 1, column, QTableWidgetItem(item))

            otdel_bled = row_items[0]
            self.table_widget9999.setItem(row_position - 1, 4, combo_2.setCurrentText(otdel_bled))

            self.table_widget9999.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
            self.table_widget9999.setHorizontalHeaderLabels(
                ["Подразделение", "Должность", "ФИО", "Статус", "Новый отдел", 'Дата',
                 'Номер приказа', 'Тип приказа', 'Дата заявления'])

            calendarr = QCalendarWidget()
            calendarr.setGridVisible(True)
            calendarr.setStyleSheet("QCalendarWidget QToolButton"
                                    "{"
                                    "background-color : lightgrey;"
                                    "color : black"
                                    "}")

            calendarchik = QtWidgets.QDateEdit(calendarPopup=True)
            calendarchik.setCalendarWidget(calendarr)
            calendarchik.setDateTime(QtCore.QDateTime.currentDateTime())
            calendarchik.setDisplayFormat('dd.MM.yyyy')
            self.table_widget9999.setCellWidget(row_position - 1, 5, calendarchik)

            calendarr2 = QCalendarWidget()
            calendarr2.setGridVisible(True)
            calendarr2.setStyleSheet("QCalendarWidget QToolButton"
                                     "{"
                                     "background-color : lightblue;"
                                     "color : black"
                                     "}")

            calendarchik2 = QtWidgets.QDateEdit(calendarPopup=True)
            calendarchik2.setCalendarWidget(calendarr2)
            calendarchik2.setDateTime(QtCore.QDateTime.currentDateTime())
            calendarchik2.setDisplayFormat('dd.MM.yyyy')
            self.table_widget9999.setCellWidget(row_position - 1, 8, calendarchik2)

            self.table_widget9999.setWordWrap(True)
            self.table_widget9999.resizeColumnsToContents()
            self.table_widget9999.setColumnWidth(0, 180)
            self.table_widget9999.setColumnWidth(1, 90)
            self.table_widget9999.setColumnWidth(2, 180)
            self.table_widget9999.setColumnWidth(5, 90)
            self.table_widget9999.resizeRowsToContents()

            log_print('Добавлено поле для ввода приказа')
        except Exception as e:
            log_print(e, 'Ошибка при добавлении поля для ввода приказа')
            pass

    def search_table(self, text):
        currentTable = self.table_widget23
        # очистка выделения
        currentTable.clearSelection()
        # поиск и обновление видимости строк
        try:
            for i in range(currentTable.rowCount()):
                matches = False
                for j in range(currentTable.columnCount()):
                    item = currentTable.item(i, j)
                    if item is not None and text.lower() in item.text().lower():
                        matches = True
                        break
                currentTable.setRowHidden(i, not matches)
            log_print('Выполнен поиск в штатке')
        except Exception as e:
            log_print(f"Ошибка, при поиске в штатке: {e}")
            print(f"Ошибка, при поиске в штатке: {e}")

    def zapoln_chelik(self):
        df2 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
        # df33 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', usecols='N')
        # df3 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A')
        df2_filtered = df2.dropna(subset=[df2.columns[1]])
        df_orders = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Приказы в работе', usecols='C')

        df2_filtered['Закрыто'] = [0 for i in range(len(df2_filtered))]
        df2_filtered.loc[df2_filtered['ФИО'].isin(df_orders['ФИО']), 'Закрыто'] = 1
        df2_filtered.loc[df2_filtered['Unnamed: 0'] == 1.000, 'Закрыто'] = 2

        df2_filtered = df2_filtered.iloc[:, 1:]

        # Установка количества строк и столбцов в QTableWidget
        self.table_widget23.setRowCount(df2_filtered.shape[0])
        self.table_widget23.setColumnCount(df2_filtered.shape[1])

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df2_filtered.shape[0]):
            for col in range(df2_filtered.shape[1]):
                xxx = str(df2_filtered.iat[row, col])
                if xxx == '0':
                    xxx = ''
                elif xxx == '1':
                    xxx = 'Заведен приказ'
                elif xxx == '2':
                    xxx = 'Должность на рассмотрении'
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget23.setItem(row, col, item)

        self.table_widget23.resizeColumnsToContents()
        self.table_widget23.setHorizontalHeaderLabels(["Подразделение", "Должность", "ФИО", "Статус приказа"])
        self.table_widget23.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        self.table_widget9999.setWordWrap(True)
        self.table_widget9999.resizeColumnsToContents()
        self.table_widget9999.setColumnWidth(0, 180)
        self.table_widget9999.setColumnWidth(1, 90)
        self.table_widget9999.setColumnWidth(2, 180)
        self.table_widget9999.setColumnWidth(5, 90)
        self.table_widget9999.resizeRowsToContents()

        self.table_widget23.setColumnWidth(3, 130)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()
        log_print('Таблица штатки успешно заполнена')


class ExcelTableWidget(QWidget):
    def __init__(self, df, parent, i):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.h_layout3 = QHBoxLayout()
        self.h_layout4 = QHBoxLayout()
        self.h_layout5 = QHBoxLayout()

        self.df = df
        self.table_widget2 = QTableWidget(self)
        self.table_widget2.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget2.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget2.setFont(font)

        self.table_widget27 = QTableWidget(self)
        self.table_widget27.setEditTriggers(QTableWidget.AllEditTriggers)
        font = QFont("Times", 10)
        self.table_widget27.setFont(font)

        self.button_45 = QPushButton("Подгрузить штат")
        self.button_45.clicked.connect(self.change_cheliki)
        self.button_45.setObjectName('blue_button')
        # self.button_45.setStyleSheet(
        #     "\n" "background-color: rgb(182, 202, 237);\n" "\n" "font: 13pt \"Times New Roman\";"
        #     "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        # self.button_45.setFixedHeight(35)

        self.button_9 = QPushButton("Добавить новую должность")
        self.button_9.clicked.connect(self.new_doljn)
        self.button_9.setObjectName('blue_button')
        # self.button_9.setStyleSheet(
        #     "\n" "background-color: rgb(182, 202, 237);\n" "\n" "font: 13pt \"Times New Roman\";"
        #     "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        # self.button_9.setFixedHeight(35)

        self.button_91 = QPushButton("Удалить новую должность")
        self.button_91.clicked.connect(self.new_doljn2)
        self.button_91.setObjectName('blue_button')
        # self.button_91.setStyleSheet(
        #     "\n" "background-color: rgb(182, 202, 237);\n" "\n" "font: 13pt \"Times New Roman\";"
        #     "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        # self.button_91.setFixedHeight(35)

        self.button_5 = QPushButton("Изменить данные сотрудников")
        self.button_5.clicked.connect(self.correct_cheliki)
        # self.button_5.setStyleSheet(
        #     "\n" "background-color: rgb(245, 193, 117);\n" "\n" "font: 13pt \"Times New Roman\";"
        #     "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        # self.button_5.setFixedHeight(35)

        self.button_59 = QPushButton("Cтатистика")
        self.button_59.clicked.connect(self.statistiks)
        self.button_59.setObjectName('blue_button')
        # self.button_59.setStyleSheet(
        #     "\n" "background-color: rgb(182, 202, 237);\n" "\n" "font: 13pt \"Times New Roman\";"
        #     "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        # self.button_59.setFixedHeight(35)

        self.combo_box = QComboBox(self)
        self.combo_box.setFixedHeight(35)
        font = QFont("Times", 10)
        self.combo_box.setFont(font)

        # self.combo_box.activated.connect(self.do_something)
        self.label105 = QLabel("Выберите из перечня необходимый отдел:")
        self.label105.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                    "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                    "font-weight: bold")

        self.h_layout3.addWidget(self.label105)
        self.h_layout3.addWidget(self.combo_box)
        self.h_layout3.addWidget(self.button_45)

        self.h_layout4.addWidget(self.button_9)
        self.h_layout4.addWidget(self.button_91)

        self.h_layout5.addWidget(self.button_59)
        self.h_layout5.addWidget(self.button_5)

        self.layout.addLayout(self.h_layout3)
        self.layout.addLayout(self.h_layout4)
        self.layout.addLayout(self.h_layout5)

        self.layout.addWidget(self.table_widget2)
        self.layout.addWidget(self.table_widget27)
        self.setLayout(self.layout)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelik()
        self.table_widget2.setHorizontalHeaderLabels(
            ["Подразделение", "Должность", "ФИО", "Статус"])
        self.table_widget2.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
        log_print('Инициализирована 2 вкладка')
        self.show()

    def statistiks(self):
        if self.table_widget27.isVisible():
            self.table_widget27.hide()
        else:
            self.table_widget27.show()

    def new_doljn(self):
        current_row_count = self.table_widget2.rowCount()
        self.table_widget2.setRowCount(current_row_count + 1)
        combo = QComboBox()
        combo.addItems(
            ["", "Должность 1", "Должность 2", "Должность 3"])
        self.table_widget2.setCellWidget(current_row_count, 3, combo)
        content = QTableWidgetItem(self.combo_box.currentText())
        self.table_widget2.setItem(current_row_count, 0, content)

    def new_doljn2(self):
        # Надо дописать удаление про выделение ячеек должностей
        current_row_count = self.table_widget2.rowCount()
        self.table_widget2.setRowCount(current_row_count - 1)

    def zapoln_chelik(self):
        df2 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', skiprows=2, usecols='M:O')
        df33 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', usecols='N')
        df000 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Статистика', usecols='P:Q')
        otdel22 = df33.iloc[0, 0]
        ############# Тут код по доп табличке########################################
        df000_filtr = df000.dropna(subset=[df000.columns[0]])
        self.table_widget27.setRowCount(df000_filtr.shape[0])
        self.table_widget27.setColumnCount(df000_filtr.shape[1] + 1)

        for row in range(df000_filtr.shape[0]):
            for col in range(df000_filtr.shape[1]):
                xxx = str(df000_filtr.iat[row, col])
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget27.setItem(row, col, item)

        self.table_widget27.resizeColumnsToContents()
        self.table_widget27.setVisible(False)
        ##########################################################################
        df2_filtered = df2.dropna(subset=[df2.columns[1]])
        # Установка количества строк и столбцов в QTableWidget
        self.table_widget2.setRowCount(df2_filtered.shape[0])
        self.table_widget2.setColumnCount(df2_filtered.shape[1] + 1)

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df2_filtered.shape[0]):
            for col in range(df2_filtered.shape[1]):
                xxx = str(df2_filtered.iat[row, col])
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget2.setItem(row, col, item)

        for eeede in range(self.table_widget2.rowCount()):
            combo = QComboBox()
            combo.addItems(
                ["", "Должность 1", "Должность 2", "Должность 3"])
            self.table_widget2.setCellWidget(eeede, 3, combo)

        self.table_widget2.resizeColumnsToContents()

        df3 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='zamruk', skiprows=3, usecols='B')
        geek_list = df3.values.tolist()
        for i in range(len(geek_list)):
            self.combo_box.addItems(geek_list[i])
        self.combo_box.setEditable(True)

        self.combo_box.setCurrentText(otdel22)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()

    def correct_cheliki(self):
        df44 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
        # Получение значения для поиска из QTableWidget
        rows = self.table_widget2.rowCount()
        content = self.combo_box.currentText()

        columnnames = ['Unnamed: 0', 'Подразделение', 'Должность', 'ФИО', 'Статус']
        dataframe6 = pd.DataFrame(columns=columnnames)
        dataframe8 = pd.DataFrame(columns=columnnames)

        for row in range(self.table_widget2.rowCount()):
            rowData = [self.table_widget2.item(row, col).text() if col != 3 else self.table_widget2.cellWidget(row,
                                                                                                               col).currentText()
                       for col in range(self.table_widget2.columnCount())]
            rowData.insert(0, " ")
            dataframe6 = pd.concat([dataframe6, pd.DataFrame([rowData], columns=columnnames)])

        a = {'Должность 1': 1, 'Должность 2': 2, 'Должность 3': 3}
        dataframe6['допжопа'] = dataframe6['Должность']
        dataframe6['допжопа'] = dataframe6['допжопа'].replace(a)

        dataframe6.sort_values(by=['допжопа', 'ФИО'], inplace=True)

        for index, row in dataframe6.iterrows():
            if row['Статус'] != '':
                if row['Статус'] == 'Уволен':
                    row['ФИО'] = np.nan
                else:
                    row['Должность'] = row['Статус']


        # Поиск первой и последней строки, содержащих 'content' в dataframe1
        first_row = df44[df44['Подразделение'] == content].index[0]
        last_row = df44[df44['Подразделение'] == content].index[-1]

        # columnnames = ['Unnamed: 0', 'Подразделение', 'Должность', 'ФИО', 'Статус']
        df55 = df44


        count_zapoln_dolgn = df55[df55.Подразделение == dataframe6.iloc[0, 1]].drop('ФИО', axis=1).replace({' ': 0}).groupby(by=['Должность']).sum().drop('Подразделение', axis=1)
        columnnames.append('допжопа')
        dataframe_with_unnamed = pd.DataFrame(columns=columnnames)

        for index, row in dataframe6.iterrows():
            comdition1 = row['Подразделение']
            comdition2 = row['Должность']
            comdition3 = row['ФИО']
            if comdition3 == '':
                if count_zapoln_dolgn.loc[comdition2][0] > 0:
                    row['Unnamed: 0'] = 1
                    count_zapoln_dolgn.loc[comdition2][0] -= 1
            dataframe_with_unnamed.loc[len(dataframe_with_unnamed.index)] = list(row)



        # Вставка dataframe6 между первой и последней строками в dataframe1

        dataframe_with_unnamed.drop(['допжопа','Статус'], axis=1, inplace=True)
        dataframe1_updated = pd.concat([df44.iloc[:first_row], dataframe_with_unnamed, df44.iloc[last_row + 1:]], axis=0)



        # Сохранение обновленного dataframe1 обратно в файл
        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Штатка']
        ws.delete_cols(1, 6)
        ws.delete_rows(1, 500)
        for r in dataframe_to_rows(dataframe1_updated, index=False, header=True):
            ws.append(r)
        # ws2 = book['Текст']
        # for r in dataframe_to_rows(dataframe8, index=False, header=True):
        #     ws2.append(r)
        book.save('addition/kadrifile.xlsx')
        book.close()

        self.message_box3 = QMessageBox(self)
        self.message_box3.setWindowTitle("Отчет об отработке")
        self.message_box3.setText("Сотрудники изменены,сейчас будет подгружен штат с внесенными изменениями, ожидайте!")
        self.message_box3.setStandardButtons(QMessageBox.Ok)
        self.message_box3.show()
        self.change_cheliki()
        self.table_widget2.setHorizontalHeaderLabels(
            ["Подразделение", "Должность", "ФИО", "Статус"])

        self.table_widget2.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

    def change_cheliki(self):
        content = self.combo_box.currentText()
        workbook = xw.Book("addition/kadrifile.xlsx")
        sheet = workbook.sheets[0]
        sheet.range(2, 14).value = content
        workbook.save()  # Сохраняем изменения в файле
        workbook.close()
        self.zapoln_chelik()
        self.message_box2 = QMessageBox(self)
        self.message_box2.setWindowTitle("Отчет об отработке")
        self.message_box2.setText("Сотрудники подгружены!")
        self.message_box2.setStandardButtons(QMessageBox.Ok)
        self.message_box2.show()

    def change_cheliki2(self):
        content = self.combo_box.currentText()
        workbook = xw.Book("addition/kadrifile.xlsx")
        sheet = workbook.sheets[0]
        sheet.range(2, 14).value = content
        workbook.save()  # Сохраняем изменения в файле
        workbook.close()
        self.zapoln_chelik2()
        self.message_box2 = QMessageBox(self)
        self.message_box2.setWindowTitle("Отчет об отработке")
        self.message_box2.setText("Сотрудники подгружены!")
        self.message_box2.setStandardButtons(QMessageBox.Ok)
        self.message_box2.show()


class ExcelTableWidget2(QWidget):
    def __init__(self, df, parent, i):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.h_layout7 = QHBoxLayout()
        self.df = df
        self.table_widget23 = QTableWidget(self)
        self.table_widget23.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget23.setStyleSheet("background-image: url(картинки/2.png)")

        font = QFont("Times", 10)
        self.table_widget23.setFont(font)

        self.label1 = QLabel("Поиск:")
        self.label1.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                  "font-weight: bold")

        self.searchLineEdit = QLineEdit()
        self.searchLineEdit.textChanged.connect(self.search_table)

        self.h_layout7.addWidget(self.label1)
        self.h_layout7.addWidget(self.searchLineEdit)

        self.layout.addLayout(self.h_layout7)
        self.layout.addWidget(self.table_widget23)
        self.setLayout(self.layout)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelik()
        self.table_widget23.setHorizontalHeaderLabels(
            ["Подразделение", "Должность", "ФИО", "Статус"])
        self.table_widget23.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        log_print('Инициализирована 3 вкладка')
        self.show()

    def search_table(self, text):
        currentTable = self.table_widget23
        # очистка выделения
        currentTable.clearSelection()
        # поиск и обновление видимости строк
        try:
            for i in range(currentTable.rowCount()):
                matches = False
                for j in range(currentTable.columnCount()):
                    item = currentTable.item(i, j)
                    if item is not None and text.lower() in item.text().lower():
                        matches = True
                        break
                currentTable.setRowHidden(i, not matches)
        except Exception as e:
            print(f"Ошибка: {e}")

    def new_doljn(self):
        current_row_count = self.table_widget23.rowCount()
        self.table_widget23.setRowCount(current_row_count + 1)
        combo = QComboBox()
        combo.addItems(
            ["", "Должность 1", "Должность 2", "Должность 3"])
        self.table_widget23.setCellWidget(current_row_count, 3, combo)

    def new_doljn2(self):
        # Надо дописать удаление про выделение ячеек должностей
        current_row_count = self.table_widget23.rowCount()
        self.table_widget23.setRowCount(current_row_count - 1)

    def zapoln_chelik(self):
        df2 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='B:D')
        df33 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', usecols='N')
        otdel22 = df33.iloc[0, 0]
        df2_filtered = df2.dropna(subset=[df2.columns[1]])

        # Установка количества строк и столбцов в QTableWidget
        self.table_widget23.setRowCount(df2_filtered.shape[0])
        self.table_widget23.setColumnCount(df2_filtered.shape[1])

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df2_filtered.shape[0]):
            for col in range(df2_filtered.shape[1]):
                xxx = str(df2_filtered.iat[row, col])
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget23.setItem(row, col, item)

        self.table_widget23.resizeColumnsToContents()
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()


class ExcelTableWidget3(QWidget):
    def __init__(self, df, parent, i):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.df = df
        self.table_widget79 = QTableWidget(self)
        self.table_widget79.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget79.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget79.setFont(font)

        self.button_910 = QPushButton("Изменить почту")
        self.button_910.clicked.connect(self.change_rassilka)
        self.button_910.setStyleSheet(
            "\n" "background-color: rgb(245, 193, 117);\n" "\n" "font: 13pt \"Times New Roman\";"
            "\n" "border: 2px solid rgb(96, 124, 173);\n" "border-radius: 10px;")
        self.button_910.setFixedHeight(35)

        self.table_widget79.setHorizontalHeaderLabels(
            ["Подразделение", "Курирующий зам.рук", "ФИО нач. отдела", "Почта", "Дата последней отправки", "Причина"])

        self.table_widget79.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
        self.layout.addWidget(self.button_910)
        self.layout.addWidget(self.table_widget79)

        self.setLayout(self.layout)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelikin()
        log_print('Инициализирована 4 вкладка')
        self.show()

    def change_rassilka(self):
        # df_history = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Историярасслылок2', skiprows=3, usecols='B:G')
        data = []
        for row in range(self.table_widget79.rowCount()):
            rowData = [self.table_widget79.item(row, 3).text()]
            data.append(rowData)
            # df_izmenpochta = pd.DataFrame([rowData])
        df6161 = pd.DataFrame(data)

        df_copy = df6161.reset_index(drop=True)
        df_values = df_copy.values.tolist()

        app = xw.App(visible=False)
        wb = xw.Book('addition/kadrifile.xlsx')
        sheet = wb.sheets('Историярасслылок2')
        sheet.range('E5').value = df_values
        wb.save('addition/kadrifile.xlsx')
        wb.close()
        app.quit()

        self.message_box3 = QMessageBox(self)
        self.message_box3.setWindowTitle("Отчет об отработке")
        self.message_box3.setText("Почты начальников отделов изменены, можете проводить рассылку")
        self.message_box3.setStandardButtons(QMessageBox.Ok)
        self.message_box3.show()

    def zapoln_chelikin(self):
        df_history = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Историярасслылок2', skiprows=3, usecols='B:G')
        # otdel22 = df33.iloc[0,0]
        ############# Тут код по доп табличке########################################
        df_history_filtr = df_history.dropna(subset=[df_history.columns[0]])
        self.table_widget79.setRowCount(df_history_filtr.shape[0])
        self.table_widget79.setColumnCount(df_history_filtr.shape[1] + 1)

        for row in range(df_history_filtr.shape[0]):
            for col in range(df_history_filtr.shape[1]):
                xxx = str(df_history_filtr.iat[row, col])
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget79.setItem(row, col, item)

        self.table_widget79.resizeColumnsToContents()
        self.table_widget79.setVisible(True)

        self.table_widget79.setHorizontalHeaderLabels(
            ["Подразделение", "Курирующий зам.рук", "ФИО нач. отдела", "Почта", "Дата последней отправки", "Причина"])
        self.table_widget79.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        self.table_widget79.resizeColumnsToContents()
        # self.layout.addWidget(self.button_910)


class ExcelTableWidget4(QWidget):
    def __init__(self, df, parent, i):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.df = df

        self.label1 = QLabel("\n"
                             "\n"
                             "\n"
                             "\n"
                             "\n"
                             "\n"
                             "'\n"
                             "По методическим вопросам: \n"
                             "По техническим вопросам:\n"
                             "1) \n"
                             "2) \n"
                             "3) \n")

        self.label1.setStyleSheet("\n" "color: rgb(38, 79, 153);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 22pt \"Open Sans Light\";\n" "\n"
                                  "font-weight: bold")

        self.pixmap1 = QPixmap('Картинки/лого3.png')
        self.laberart = QLabel(self)
        self.laberart.setPixmap(self.pixmap1)
        self.laberart.setGeometry(QtCore.QRect(10, 10, 350, 450))

        # self.layout.addWidget(self.laberart)
        self.layout.addWidget(self.label1)
        self.setLayout(self.layout)

        log_print('Инициализирована 5 вкладка')
        self.show()


class ExcelTableWidget5(QMainWindow):
    def __init__(self, df, parent, i):
        # эта штука для создания дашбордов
        # поясню на примере примере ПЕРВОГО дашборда, чтобы не повторяться(так как они схожие)

        super().__init__(parent)
        self.setWindowTitle("Анимированный дашборд")
        self.setWindowTitle("Creating Barchart")
        self.setWindowIcon(QIcon("python.png"))  # хз нахрена это сюда вставили но пусть будет
        self.setStyleSheet('background-color:rgb(31,32,40)')  # цвет фона указали

        # создали кароч окно  и надо добавить лейаут для общей отрисовки дащбордов в них
        self.scroll = QScrollArea()  # Scroll Area which contains the widgets, set as the centralWidget
        self.widget = QWidget()
        self.layout = QHBoxLayout(self)
        self.layout_MAIN = QVBoxLayout(self)
        self.widget.setLayout(self.layout_MAIN)

        # тут я кароч забираю данные из эксельки для того чтобы верхняя плашка была заполнена данными по отдельным графам
        df_infos = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='I:K')
        df_infos2 = df_infos.dropna(subset=[df_infos.columns[0]])
        shtat1 = df_infos2.iloc[0, 1]
        shtat1 = int(shtat1)
        shtat2 = df_infos2.iloc[1, 0]
        shtat2 = int(shtat2)
        shtat3 = df_infos2.iloc[2, 0]
        shtat3 = int(shtat3)
        shtat4 = df_infos2.iloc[0, 0]
        shtat4 = int(shtat4)
        shtat5 = df_infos2.iloc[0, 2]
        shtat5 = int(shtat5)
        datenowpokazat = df_infos2.iloc[3, 0]
        datenowpokazat = str(datenowpokazat)

        # общее название странички
        self.labelnaim = QLabel("Штатная численность МБУ ФК", self)
        self.labelnaim.setAcceptDrops(True)
        self.labelnaim.setAlignment(Qt.AlignCenter)
        self.labelnaim.setAutoFillBackground(False)
        self.labelnaim.setScaledContents(True)
        self.labelnaim.setWordWrap(True)
        self.labelnaim.setStyleSheet("\n" "color: rgb(57, 125, 214);\n"
                                     "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 27pt \"Myriad pro\";\n" "\n"
                                     "font-weight: bold")
        # дата ласт обновл подтягивается тоже из эксельки суть в том что после обновл другими, он попадает в общий
        self.labeltimeobn = QLabel(f"Актуальность дашборда на\n {datenowpokazat}", self)
        self.labeltimeobn.setAcceptDrops(True)
        self.labeltimeobn.setAlignment(Qt.AlignCenter)
        self.labeltimeobn.setAutoFillBackground(False)
        self.labeltimeobn.setScaledContents(True)
        self.labeltimeobn.setWordWrap(True)
        self.labeltimeobn.setStyleSheet("\n" "color: rgb(107, 107, 107);\n"
                                        "\n" "background-color: rgb(107, 107, 107,0);\n" "\n" "font: 15pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")
        # начинается пиздец
        # для каждой пары подпись-цифра создается свой лейаут в котором по номеру видну какой лейаут какой лейбл и что отвечает
        # за цифру тип наименовние имеет индекс1 поэтому в конце лейаута наим и цифры подписи индекс вторая цифра порядок сверху
        # те наименование точно 1-2-3 цифра имеет подпист 12-22-32
        self.layout_naim1 = QVBoxLayout()
        self.labelnaime1 = QLabel("Штатная численность", self)
        self.labelnaime1.setAcceptDrops(True)
        self.labelnaime1.setAutoFillBackground(False)
        self.labelnaime1.setAlignment(Qt.AlignCenter)
        self.labelnaime1.setScaledContents(True)
        self.labelnaime1.setWordWrap(True)
        self.labelnaime1.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")
        self.labelnaime12 = QLabel(f"{shtat1}", self)
        self.labelnaime12.setAcceptDrops(True)
        self.labelnaime12.setAutoFillBackground(False)
        self.labelnaime12.setAlignment(Qt.AlignCenter)
        self.labelnaime12.setScaledContents(True)
        self.labelnaime12.setWordWrap(True)
        self.labelnaime12.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")
        self.layout_naim1.addWidget(self.labelnaime1)
        self.layout_naim1.addWidget(self.labelnaime12)

        self.layout_naim2 = QVBoxLayout()
        self.labelnaime2 = QLabel("Фактическое количество сотрудников \n(с учетом декретных должн.)", self)
        self.labelnaime2.setAcceptDrops(True)
        self.labelnaime2.setAutoFillBackground(False)
        self.labelnaime2.setAlignment(Qt.AlignCenter)
        self.labelnaime2.setScaledContents(True)
        self.labelnaime2.setWordWrap(True)
        self.labelnaime2.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")

        self.layoutinlabelname2 = QHBoxLayout()
        self.labelnaime22 = QLabel(f"{shtat2}", self)
        self.labelnaime22.setAcceptDrops(True)
        self.labelnaime22.setAutoFillBackground(False)
        self.labelnaime22.setAlignment(Qt.AlignRight)
        self.labelnaime22.setScaledContents(True)
        self.labelnaime22.setWordWrap(True)
        self.labelnaime22.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")
        self.labelnaime23 = QLabel(f"     {int((shtat2 / shtat1) * 100)}%", self)
        self.labelnaime23.setAcceptDrops(True)
        self.labelnaime23.setAutoFillBackground(False)
        self.labelnaime23.setAlignment(Qt.AlignLeft)
        self.labelnaime23.setScaledContents(True)
        self.labelnaime23.setWordWrap(True)
        self.labelnaime23.setStyleSheet("\n" "color: rgb(255, 159, 0);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")

        self.layoutinlabelname2.addWidget(self.labelnaime22)
        self.layoutinlabelname2.addWidget(self.labelnaime23)
        self.layout_naim2.addWidget(self.labelnaime2)
        self.layout_naim2.addLayout(self.layoutinlabelname2)

        self.layout_naim3 = QVBoxLayout()
        self.labelnaime3 = QLabel("Количество сотрудников \nв декрете", self)
        self.labelnaime3.setAcceptDrops(True)
        self.labelnaime3.setAutoFillBackground(False)
        self.labelnaime3.setAlignment(Qt.AlignCenter)
        self.labelnaime3.setScaledContents(True)
        self.labelnaime3.setWordWrap(True)
        self.labelnaime3.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")

        self.labelnaime32 = QLabel(f"{shtat3}", self)
        self.labelnaime32.setAcceptDrops(True)
        self.labelnaime32.setAutoFillBackground(False)
        self.labelnaime32.setAlignment(Qt.AlignCenter)
        self.labelnaime32.setScaledContents(True)
        self.labelnaime32.setWordWrap(True)
        self.labelnaime32.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")
        self.layout_naim3.addWidget(self.labelnaime3)
        self.layout_naim3.addWidget(self.labelnaime32)

        self.layout_naim4 = QVBoxLayout()
        self.labelnaime4 = QLabel("Фактическое количество \nсотрудников", self)
        self.labelnaime4.setAcceptDrops(True)
        self.labelnaime4.setAutoFillBackground(False)
        self.labelnaime4.setAlignment(Qt.AlignCenter)
        self.labelnaime4.setScaledContents(True)
        self.labelnaime4.setWordWrap(True)
        self.labelnaime4.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")
        self.labelnaime42 = QLabel(f"{shtat4}", self)
        self.labelnaime42.setAcceptDrops(True)
        self.labelnaime42.setAutoFillBackground(False)
        self.labelnaime42.setAlignment(Qt.AlignCenter)
        self.labelnaime42.setScaledContents(True)
        self.labelnaime42.setWordWrap(True)
        self.labelnaime42.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")

        self.layout_naim4.addWidget(self.labelnaime4)
        self.layout_naim4.addWidget(self.labelnaime42)

        self.layout_naim5 = QVBoxLayout()
        self.labelnaime5 = QLabel("Количество вакантных мест \n(с учетом декретных долж.)", self)
        self.labelnaime5.setAcceptDrops(True)
        self.labelnaime5.setAutoFillBackground(False)
        self.labelnaime5.setScaledContents(True)
        self.labelnaime5.setAlignment(Qt.AlignCenter)
        self.labelnaime5.setWordWrap(True)
        self.labelnaime5.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")
        self.layoutinlabelname5 = QHBoxLayout()
        self.labelnaime52 = QLabel(f"{shtat5}", self)
        self.labelnaime52.setAcceptDrops(True)
        self.labelnaime52.setAutoFillBackground(False)
        self.labelnaime52.setScaledContents(True)
        self.labelnaime52.setAlignment(Qt.AlignRight)
        self.labelnaime52.setWordWrap(True)
        self.labelnaime52.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")
        self.labelnaime53 = QLabel(f"     {int((shtat5 / shtat1) * 100)}%", self)
        self.labelnaime53.setAcceptDrops(True)
        self.labelnaime53.setAutoFillBackground(False)
        self.labelnaime53.setAlignment(Qt.AlignLeft)
        self.labelnaime53.setScaledContents(True)
        self.labelnaime53.setWordWrap(True)
        self.labelnaime53.setStyleSheet("\n" "color: rgb(255, 159, 0);\n"
                                        "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 18pt \"Myriad pro\";\n" "\n"
                                        "font-weight: bold")

        self.layoutinlabelname5.addWidget(self.labelnaime52)
        self.layoutinlabelname5.addWidget(self.labelnaime53)
        self.layout_naim5.addWidget(self.labelnaime5)
        self.layout_naim5.addLayout(self.layoutinlabelname5)

        self.layout_naim6 = QVBoxLayout()
        self.labelnaime6 = QLabel("В процессе \nтрудоустройства", self)
        self.labelnaime6.setAcceptDrops(True)
        self.labelnaime6.setAutoFillBackground(False)
        self.labelnaime6.setScaledContents(True)
        self.labelnaime6.setAlignment(Qt.AlignCenter)
        self.labelnaime6.setWordWrap(True)
        self.labelnaime6.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                       "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 10pt \"Myriad pro\";\n" "\n"
                                       "font-weight: bold")

        self.layout_naim6.addWidget(self.labelnaime6)

        # добавляем в общий лейаут вот эти сверху созданные хуйни
        self.first_layout = QVBoxLayout()
        self.first_layout.addWidget(self.labelnaim)
        self.first_layout.addWidget(self.labeltimeobn)
        self.second_layout = QHBoxLayout()
        self.second_layout.addLayout(self.layout_naim1)
        self.second_layout.addLayout(self.layout_naim2)
        self.second_layout.addLayout(self.layout_naim3)
        self.second_layout.addLayout(self.layout_naim4)
        self.second_layout.addLayout(self.layout_naim5)
        self.second_layout.addLayout(self.layout_naim6)

        # тут создаем 2 кнопеи по сбросу фильтров и обновл дашборда
        self.second_layoutknopks = QHBoxLayout()
        self.pushButtonudalfiltr = QPushButton("Сбросить все фильтры", self)
        self.pushButtonudalfiltr.setFixedHeight(30)
        self.pushButtonudalfiltr.clicked.connect(self.vosstanovit)
        self.pushButtonudalfiltr.setStyleSheet("color: rgb(255, 255, 255);\n"
                                               "\n" "background-color: rgb(57, 125, 214);\n" "\n" "font: 12pt \"Myriad pro\";"
                                               "\n" "border: 4px solid rgb(6, 73, 129);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.pushButtonudalfiltr2 = QPushButton("Обновить дашборд", self)
        self.pushButtonudalfiltr2.setFixedHeight(30)
        self.pushButtonudalfiltr2.clicked.connect(self.renewdash)
        self.pushButtonudalfiltr2.setStyleSheet("color: rgb(255, 255, 255);\n"
                                                "\n" "background-color: rgb(57, 125, 214);\n" "\n" "font: 12pt \"Myriad pro\";"
                                                "\n" "border: 3px solid rgb(6, 73, 129);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.second_layoutknopks.addWidget(self.pushButtonudalfiltr)
        self.second_layoutknopks.addWidget(self.pushButtonudalfiltr2)
        # опять закинули в общий лейаут
        self.layout_MAIN.addLayout(self.first_layout)
        self.layout_MAIN.addLayout(self.second_layout)
        self.layout_MAIN.addLayout(self.second_layoutknopks)
        ############################# Первый дэш ЗАМРУКИ
        # забираем из эксельки датафрейм для отрисовки и фильтруем чтобы убрать nan из листа эксельки
        dfcheckk = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='W:AA')
        df2_checkk = dfcheckk.dropna(subset=[dfcheckk.columns[1]])
        filtered_df = df2_checkk.sort_values(by=df2_checkk.columns[0], ascending=False)

        # тут создаем 3 графы которые мы будем видеть первые две будут спаянными третья попадет в правую часть как отдельный вид
        # отдельно потому что нельзя его встроить тк пкьютчарт говно говна в котором нельзя к одной графе применить проценты или инт
        self.set01 = QBarSet("Факт")
        self.set11 = QBarSet("Вакант")
        self.set21 = QBarSet("% Ваканта")

        # тут создаем два листа для того чтобы в будущем отобразить их справа и слева от самого графика(подписи будут)
        vakant1 = []
        zamrukk = []

        # тут перебираем все строки в датафрейме через индекс роу роу содержит строку целиком поэтому если в скобках
        # квардратных написать цифру то обращаемся к конкретному столбцу в этой строке
        # далее начинаем с конца за это отвечает -1 это потому что при добавлении в значение он добавляет как видит
        # те если хотим чтобы алфавитно было а в дашборде у нас уже алфавитно тупо пишем-1
        # еще он тут че делает он тут кароч кроме добавления так же и отвечает за проценты(правая часть табл
        # )тут очень внимательно так как если у вас будут одинаковые проценты он тупо скроет значение какое-то
        # поэтому в некоторых примерах есть индекс который мы добавили чтобы при отрисовке он показывал все варианты
        for index, row in filtered_df[::-1].iterrows():
            label = row[0]
            zamrukk.append(label)
            value1 = int(row[1])
            value2 = int(row[3])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            value3 = str(okrugl) + "%"
            self.set01.append(value1)
            self.set11.append(value2)
            vakant1.append(value3)

        # ВСЕ ПОГНАЛИ БЛЯТЬ
        # тут в чем суть он создает стак бар серис это тип один граф спаивается с другим если нужны другие графики то в документацию
        # то есть мы объявили что сериес это у нас спаянный дальше в него добавляем через аппенд 2 значения который мы раннее посчитали
        # последние 2 строки отвечают за подписи все стандартно подписи тру и подписи в конце
        self.series1 = QHorizontalStackedBarSeries()
        self.series1.append(self.set01)
        self.series1.append(self.set11)
        self.series1.setLabelsVisible(True)
        self.series1.setLabelsPosition(QBarSeries.LabelsInsideEnd)

        # создаем сам график в него закидываем сериес в котором мы так хуйнюли значения до этого задаем наименование и тип фона графика
        self.chart1 = QChart()
        self.chart1.addSeries(self.series1)
        self.chart1.setTitle("Количество сотрудников согласно списку Заместителей Руководителя")
        self.chart1.setAnimationOptions(QChart.SeriesAnimations)
        self.chart1.setTheme(QChart.ChartThemeDark)

        # создаем правые подписи так как (Y-2), заполняем из ранее указанного листа вакант ТК ЭТО ПРОЦЕНТ даем им наименование + оформление
        self.axisY12 = QBarCategoryAxis()
        self.axisY12.append(vakant1)
        self.axisY12.setTitleText("Процент Ваканта от общего числа")
        self.axisY12.setTitleBrush(QColor('#ff9f00'))
        self.axisY12.setLabelsFont(QFont("Times", 10))
        self.axisY12.setLabelsColor(QColor('#ff9f00'))
        self.chart1.addAxis(self.axisY12, Qt.AlignRight)
        self.series1.attachAxis(self.axisY12)

        # тоже самое что и раньше только уже левые подписи тк (Y1-)
        self.axisY1 = QBarCategoryAxis()
        self.axisY1.append(zamrukk)
        self.axisY1.setLabelsFont(QFont("Times", 11))
        self.chart1.addAxis(self.axisY1, Qt.AlignLeft)
        self.series1.attachAxis(self.axisY1)

        # подписи для Х
        self.axisX1 = QValueAxis()
        self.axisX1.setLabelsFont(QFont("Times", 11))
        self.chart1.addAxis(self.axisX1, Qt.AlignBottom)
        self.series1.attachAxis(self.axisX1)

        # создаем отображение графика и задаем его фиксированные размеры
        self.chartview1 = QChartView(self.chart1)
        self.chartview1.setRenderHint(QPainter.Antialiasing)
        self.chartview1.setFixedSize(QSize(850, 480))

        # Создаем часть с фильтрами
        self.label3331 = QLabel()
        self.label3331.setText("Фильтр")
        self.label3331.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                     "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 14pt \"Open Sans Light\";\n" "\n"
                                     "font-weight: bold")

        self.pushButton216811 = QPushButton("По факт", self)
        self.pushButton216811.clicked.connect(lambda: self.click_filtrzamruk(1))
        self.pushButton216811.setFixedHeight(25)
        self.pushButton216811.setFixedWidth(100)
        self.pushButton216811.setStyleSheet("color: rgb(59, 59, 59);\n"
                                            "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                            "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.pushButton216812 = QPushButton("Общее число", self)
        self.pushButton216812.clicked.connect(lambda: self.click_filtrzamruk(2))
        self.pushButton216812.setFixedHeight(25)
        self.pushButton216812.setFixedWidth(100)
        self.pushButton216812.setStyleSheet("color: rgb(59, 59, 59);\n"
                                            "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                            "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.pushButton216813 = QPushButton("По вакант", self)
        self.pushButton216813.clicked.connect(lambda: self.click_filtrzamruk(3))
        self.pushButton216813.setFixedHeight(25)
        self.pushButton216813.setFixedWidth(100)
        self.pushButton216813.setStyleSheet("color: rgb(59, 59, 59);\n"
                                            "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                            "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        # тут я кароч создаю вызов функции по отбору в других дашборд после клика по конкретному столбику
        self.series1.clicked.connect(self.bar_double_clicked)

        self.v_layoutfiltr = QVBoxLayout()
        self.v_layoutfiltr.addWidget(self.label3331)
        self.v_layoutfiltr.addWidget(self.pushButton216811)
        self.v_layoutfiltr.addWidget(self.pushButton216812)
        self.v_layoutfiltr.addWidget(self.pushButton216813)
        self.layout.addLayout(self.v_layoutfiltr)
        self.layout.addWidget(self.chartview1)

        # здесь задаем цвет шрифт и его размер для именно столбиков которые мы отрисовали
        self.set01.setColor(QColor("#397dd6"))
        self.set01.setLabelFont(QFont("Times", 11))

        self.set11.setColor(QColor("#b2ebf2"))
        self.set11.setLabelColor(QColor("#616161"))
        self.set11.setLabelFont(QFont("Times", 11))

        self.set21.setLabelColor(QColor("#616161"))
        self.set21.setLabelFont(QFont("Times", 11))

        ################################# второй дэш
        dfcheckk2 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='A:G')
        df2_checkk2 = dfcheckk2.dropna(subset=[dfcheckk2.columns[1]])
        grouped_df = df2_checkk2.groupby(by=df2_checkk2.columns[6]).sum()

        self.set02 = QBarSet("Факт")
        self.set12 = QBarSet("Вакант")
        self.set22 = QBarSet("% Ваканта")
        vakant2 = []
        zamrukk = []
        for index, row in grouped_df[::-1].iterrows():
            zamrukk.append(index)
            value1 = int(row[3])
            value2 = int(row[5])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            value3 = str(okrugl) + "%"
            self.set02.append(value1)
            self.set12.append(value2)
            vakant2.append(value3)

        self.series2 = QHorizontalStackedBarSeries()
        self.series2.append(self.set02)
        self.series2.append(self.set12)
        self.series2.setLabelsVisible(True)
        self.series2.setLabelsPosition(QBarSeries.LabelsInsideEnd)

        self.chart2 = QChart()
        self.chart2.addSeries(self.series2)
        self.chart2.setTitle("Количество сотрудников согласно должностям")
        self.chart2.setAnimationOptions(QChart.SeriesAnimations)
        self.chart2.setTheme(QChart.ChartThemeDark)

        self.axisY2 = QBarCategoryAxis()
        self.axisY2.append(zamrukk)
        self.axisY2.setLabelsFont(QFont("Times", 10))
        self.chart2.addAxis(self.axisY2, Qt.AlignLeft)
        self.series2.attachAxis(self.axisY2)

        self.axisY22 = QBarCategoryAxis()
        self.axisY22.append(vakant2)
        self.axisY22.setTitleText("Процент Ваканта от общего числа")
        self.axisY22.setTitleBrush(QColor('#ff9f00'))
        self.axisY22.setLabelsFont(QFont("Times", 9))
        self.axisY22.setLabelsColor(QColor('#ff9f00'))
        self.chart2.addAxis(self.axisY22, Qt.AlignRight)
        self.series2.attachAxis(self.axisY22)

        self.axisX2 = QValueAxis()
        self.axisX2.setLabelsFont(QFont("Times", 9))
        self.chart2.addAxis(self.axisX2, Qt.AlignBottom)
        self.series2.attachAxis(self.axisX2)

        self.chartview2 = QChartView(self.chart2)
        self.chartview2.setFixedSize(QSize(945, 480))

        self.layout.addWidget(self.chartview2)

        self.set02.setColor(QColor("#397dd6"))
        self.set02.setLabelFont(QFont("Times", 11))

        self.set12.setColor(QColor("#b2ebf2"))
        self.set12.setLabelColor(QColor("#616161"))
        self.set12.setLabelFont(QFont("Times", 11))

        self.set22.setLabelColor(QColor("#616161"))
        self.set22.setLabelFont(QFont("Times", 11))

        ##################################### дашборд по отделам
        dfcheckk3 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='O:S')
        df_checkk3 = dfcheckk3.dropna(subset=[dfcheckk3.columns[1]])

        self.set03 = QBarSet("Факт")
        self.set13 = QBarSet("Вакант")
        self.set23 = QBarSet("% Ваканта")
        vakant3 = []
        zamrukk = []
        for index, row in df_checkk3[::-1].iterrows():
            label = row[0]
            zamrukk.append(label)
            value1 = int(row[1])
            value2 = int(row[3])
            okrugl = round((value2 / (value1 + value2) * 100), 1)
            value3 = str(okrugl) + str(index) + "%"
            self.set03.append(value1)
            self.set13.append(value2)
            vakant3.append(value3)

        self.series3 = QHorizontalStackedBarSeries()
        self.series3.append(self.set03)
        self.series3.append(self.set13)
        self.series3.setLabelsVisible(True)
        self.series3.setLabelsPosition(QBarSeries.LabelsInsideEnd)

        self.chart3 = QChart()
        self.chart3.addSeries(self.series3)
        self.chart3.setAnimationOptions(QChart.SeriesAnimations)
        self.chart3.setTheme(QChart.ChartThemeDark)

        self.axisY3 = QBarCategoryAxis()
        self.axisY3.append(zamrukk)
        self.axisY3.setLabelsFont(QFont("Times", 8))
        self.chart3.addAxis(self.axisY3, Qt.AlignLeft)
        self.series3.attachAxis(self.axisY3)

        self.axisY32 = QBarCategoryAxis()
        self.axisY32.append(vakant3)
        self.axisY32.setTitleText("Процент Ваканта от общего числа")
        self.axisY32.setTitleBrush(QColor('#ff9f00'))
        self.axisY32.setLabelsFont(QFont("Times", 8))
        self.axisY32.setLabelsColor(QColor('#ff9f00'))
        self.chart3.addAxis(self.axisY32, Qt.AlignRight)
        self.series3.attachAxis(self.axisY32)

        self.axisX3 = QValueAxis()
        self.axisX3.setLabelsFont(QFont("Times", 8))
        self.series3.attachAxis(self.axisX3)

        self.chartview3 = QChartView(self.chart3)
        self.chartview3.setFixedSize(QSize(1780, 1050))  # Задайте желаемые размеры графика здесь
        self.chart3.setAnimationOptions(QChart.SeriesAnimations)
        self.chart3.setTheme(QChart.ChartThemeDark)

        self.label3331 = QLabel()
        self.label3331.setText("Фильтр")
        self.label3331.setStyleSheet("\n" "color: rgb(255, 255, 255);\n"
                                     "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 14pt \"Open Sans Light\";\n" "\n"
                                     "font-weight: bold")

        self.pushButton2168113 = QPushButton("По факт", self)
        self.pushButton2168113.clicked.connect(lambda: self.click_filtrotdel(1))
        self.pushButton2168113.setFixedHeight(25)
        self.pushButton2168113.setFixedWidth(100)
        self.pushButton2168113.setStyleSheet("color: rgb(59, 59, 59);\n"
                                             "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                             "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.pushButton2168123 = QPushButton("Общ число", self)
        self.pushButton2168123.clicked.connect(lambda: self.click_filtrotdel(2))
        self.pushButton2168123.setFixedHeight(25)
        self.pushButton2168123.setFixedWidth(100)
        self.pushButton2168123.setStyleSheet("color: rgb(59, 59, 59);\n"
                                             "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                             "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.pushButton2168133 = QPushButton("По вакант", self)
        self.pushButton2168133.clicked.connect(lambda: self.click_filtrotdel(3))
        self.pushButton2168133.setFixedHeight(25)
        self.pushButton2168133.setFixedWidth(100)
        self.pushButton2168133.setStyleSheet("color: rgb(59, 59, 59);\n"
                                             "\n" "background-color: rgb(255, 159, 0);\n" "\n" "font: 10pt \"Myriad\";"
                                             "\n" "border: 3px solid rgb(96, 124, 173);\n" "border-radius: 10px;" "\n" "font-weight: bold;\n")

        self.v_layoutfiltr3 = QVBoxLayout()
        self.v_layoutfiltr3.addWidget(self.label3331)
        self.v_layoutfiltr3.addWidget(self.pushButton2168113)
        self.v_layoutfiltr3.addWidget(self.pushButton2168123)
        self.v_layoutfiltr3.addWidget(self.pushButton2168133)
        vbox3 = QHBoxLayout()
        vbox3.addLayout(self.v_layoutfiltr3)
        vbox3.addWidget(self.chartview3)

        self.set03.setColor(QColor("#397dd6"))
        self.set03.setLabelFont(QFont("Times", 11))

        self.set13.setColor(QColor("#b2ebf2"))
        self.set13.setLabelColor(QColor("#616161"))
        self.set13.setLabelFont(QFont("Times", 11))

        self.set23.setLabelColor(QColor("#616161"))
        self.set23.setLabelFont(QFont("Times", 11))

        self.series3.clicked.connect(self.bar_double_clicked23)

        self.layout_MAIN.addLayout(self.layout)
        self.layout_MAIN.addLayout(vbox3)
        ##################################### дашборд по беременным
        dfcheckk4 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Штатка', usecols='B:G')
        df_checkk4 = dfcheckk4.dropna(subset=[dfcheckk4.columns[4]])
        dfcheckk44 = df_checkk4.groupby(by=df_checkk4.columns[0]).sum()
        dfcheckk44[dfcheckk44.columns[4]] = dfcheckk44[dfcheckk44.columns[4]].replace(np.nan, 0)
        dfcheckk44[dfcheckk44.columns[4]] = dfcheckk44[dfcheckk44.columns[4]].replace(2, 1)

        self.set04 = QBarSet("Сотрудники в декрете")
        self.set14 = QBarSet("Занятые должности")

        zamrukk = []
        for index, row in dfcheckk44[::-1].iterrows():
            index2 = index
            zamrukk.append(index2)
            value1 = int(row[3])
            value2 = int(row[4])
            self.set04.append(value1)
            self.set14.append(value2)

        self.series42 = QHorizontalBarSeries()
        self.series42.append(self.set14)
        self.series42.append(self.set04)
        self.series42.setLabelsVisible(True)
        self.series42.setLabelsPosition(QBarSeries.LabelsInsideEnd)

        self.chart4 = QChart()
        self.chart4.addSeries(self.series42)
        self.chart4.setAnimationOptions(QChart.SeriesAnimations)
        self.chart4.setTheme(QChart.ChartThemeDark)

        self.axisY4 = QBarCategoryAxis()
        self.axisY4.append(zamrukk)
        self.axisY4.setLabelsFont(QFont("Times", 10))
        self.chart4.addAxis(self.axisY4, Qt.AlignLeft)
        self.series42.attachAxis(self.axisY4)

        self.axisX4 = QValueAxis()
        self.axisX4.setLabelsFont(QFont("Times", 8))
        self.chart4.addAxis(self.axisX4, Qt.AlignBottom)
        self.series42.attachAxis(self.axisX4)

        self.chartview4 = QChartView(self.chart4)
        self.chartview4.setFixedSize(QSize(1500, 800))  # Задайте желаемые размеры графика здесь
        self.chart4.setAnimationOptions(QChart.SeriesAnimations)
        self.chart4.setTheme(QChart.ChartThemeDark)

        vbox4 = QHBoxLayout()
        vbox4.addWidget(self.chartview4)

        self.set04.setColor(QColor("#c8e6c9"))
        self.set04.setLabelFont(QFont("Times", 12))
        self.set04.setLabelColor(QColor("#212121"))

        self.set14.setColor(QColor("#c5cae9"))
        self.set14.setLabelColor(QColor("#212121"))
        self.set14.setLabelFont(QFont("Times", 12))

        self.chart4.setTitle("Количество сотрудников в декрете + занятые")

        self.layout_MAIN.addLayout(self.layout)
        self.layout_MAIN.addLayout(vbox4)

        # тут по скроллингу нижний отрублен но если надо то можно и врубить
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(self.widget)

        self.setCentralWidget(self.scroll)

    def bar_double_clicked23(self,
                             index):  # эта хуйня для того чтобы можно было по нажатому сделать фильтр и вывести другие данные в других дашб
        # забрали значение нажатого столбика
        otdels = self.axisY3.categories()[index]
        otdels2 = otdels

        # очистили графы чтобы новые записи потом отразить
        self.series2.clear()
        self.axisY2.clear()
        self.axisY22.clear()

        dfcheckk2 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='A:G')
        df2_checkk2 = dfcheckk2.dropna(subset=[dfcheckk2.columns[1]])
        df2_checkk27 = df2_checkk2.loc[df2_checkk2.iloc[:, 1] == otdels]
        grouped_df = df2_checkk27.groupby(by=df2_checkk27.columns[6]).sum()

        self.set02 = QBarSet("Факт")
        self.set12 = QBarSet("Вакант")
        self.set22 = QBarSet("% Ваканта")
        vakant2 = []
        otdels = []
        kolich = 0
        for index, row in grouped_df[::-1].iterrows():
            otdels.append(index)
            value1 = int(row[3])
            value2 = int(row[5])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            kolich += 1
            value3 = str(okrugl) + str(kolich) + "%"
            self.set02.append(value1)
            self.set12.append(value2)
            vakant2.append(value3)

        self.series2.append(self.set02)
        self.series2.append(self.set12)

        self.axisY2.append(otdels)
        self.axisY22.append(vakant2)

        # self.series2.attachAxis(self.axisX2)
        # self.chart2.addAxis(self.axisX2, Qt.AlignBottom)

        self.chart2.addAxis(self.axisY22, Qt.AlignRight)
        self.series2.attachAxis(self.axisY22)

        self.set02.setColor(QColor("#397dd6"))
        self.set02.setLabelFont(QFont("Times", 11))

        self.set12.setColor(QColor("#b2ebf2"))
        self.set12.setLabelColor(QColor("#616161"))
        self.set12.setLabelFont(QFont("Times", 11))

        self.set22.setLabelColor(QColor("#616161"))
        self.set22.setLabelFont(QFont("Times", 11))

        self.chart2.setTitle(f"{otdels2}")

    def bar_double_clicked(self, index, barset):
        zamrukkkk = self.axisY1.categories()[index]
        zamrukkkk2 = zamrukkkk

        self.series3.clear()
        self.axisY3.clear()
        self.axisY32.clear()

        dfcheckk3 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='N:S')
        df_checkk3 = dfcheckk3.dropna(subset=[dfcheckk3.columns[1]])
        df_checkk32 = df_checkk3.loc[df_checkk3.iloc[:, 0] == zamrukkkk]

        self.set03 = QBarSet("Факт")
        self.set13 = QBarSet("Вакант")
        self.set23 = QBarSet("% Ваканта")
        zamrukk = []
        vakant3 = []
        for index, row in df_checkk32[::-1].iterrows():
            label = row[1]
            zamrukk.append(label)
            value1 = int(row[2])
            value2 = int(row[4])
            okrugl = round((value2 / (value1 + value2) * 100), 1)
            value3 = str(okrugl) + str(index) + "%"
            self.set03.append(value1)
            self.set13.append(value2)
            vakant3.append(value3)

        self.series3.append(self.set03)
        self.series3.append(self.set13)

        self.axisY3.append(zamrukk)
        self.axisY32.append(vakant3)

        self.axisY3.setLabelsFont(QFont("Times", 8))

        self.chart3.addAxis(self.axisY3, Qt.AlignLeft)
        self.series3.attachAxis(self.axisY3)

        self.set03.setColor(QColor("#397dd6"))
        self.set03.setLabelFont(QFont("Times", 11))

        self.set13.setColor(QColor("#b2ebf2"))
        self.set13.setLabelColor(QColor("#616161"))
        self.set13.setLabelFont(QFont("Times", 11))

        self.set23.setLabelColor(QColor("#616161"))
        self.set23.setLabelFont(QFont("Times", 11))

        self.series2.clear()
        self.axisY2.clear()
        self.axisY22.clear()

        dfcheckk2 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='A:G')
        df2_checkk2 = dfcheckk2.dropna(subset=[dfcheckk2.columns[1]])
        df2_checkk27 = df2_checkk2.loc[df2_checkk2.iloc[:, 0] == zamrukkkk]
        grouped_df = df2_checkk27.groupby(by=df2_checkk27.columns[6]).sum()

        self.set02 = QBarSet("Факт")
        self.set12 = QBarSet("Вакант")
        self.set22 = QBarSet("% Ваканта")
        vakant2 = []
        zamrukk = []
        kolich = 0
        for index, row in grouped_df[::-1].iterrows():
            zamrukk.append(index)
            value1 = int(row[3])
            value2 = int(row[5])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            kolich += 1
            value3 = str(okrugl) + str(kolich) + "%"
            self.set02.append(value1)
            self.set12.append(value2)
            vakant2.append(value3)

        self.series2.append(self.set02)
        self.series2.append(self.set12)

        self.axisY2.append(zamrukk)
        self.axisY22.append(vakant2)

        self.series2.attachAxis(self.axisY2)

        self.chart2.addAxis(self.axisY22, Qt.AlignRight)
        self.series2.attachAxis(self.axisY22)

        self.series2.attachAxis(self.axisX2)

        self.set02.setColor(QColor("#397dd6"))
        self.set02.setLabelFont(QFont("Times", 10))

        self.set12.setColor(QColor("#b2ebf2"))
        self.set12.setLabelColor(QColor("#616161"))
        self.set12.setLabelFont(QFont("Times", 10))

        self.set22.setLabelColor(QColor("#616161"))
        self.set22.setLabelFont(QFont("Times", 10))

        self.chart3.setTitle(f"{zamrukkkk2}")
        self.chart2.setTitle(f"{zamrukkkk2}")

    # def clickedBarSet(self, index, _set):
    #     print(index, _set)
    #     if _set.color().name() == "#800080":
    #         _set.setColor(QColor("#008080"))
    #     else:
    #         _set.setColor(QColor("#800080"))

    def click_filtrzamruk(self, int_xxx):  # хуйня по фильтрам первого чтобы норм фильтровал значения
        self.series1.clear()
        self.axisY1.clear()
        self.axisY12.clear()
        dfcheckk = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='W:AA')
        df2_checkk = dfcheckk.dropna(subset=[dfcheckk.columns[1]])
        filtered_df = df2_checkk.sort_values(by=df2_checkk.columns[int_xxx], ascending=False)

        self.set01 = QBarSet("Факт")
        self.set11 = QBarSet("Вакант")
        self.set21 = QBarSet("% Ваканта")
        zamrukk = []
        vakant1 = []
        for index, row in filtered_df[::-1].iterrows():
            label = row[0]
            zamrukk.append(label)
            value1 = int(row[1])
            value2 = int(row[3])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            value3 = str(okrugl) + "%"
            self.set01.append(value1)
            self.set11.append(value2)
            vakant1.append(value3)

        self.series1.append(self.set01)
        self.series1.append(self.set11)

        self.axisY1.append(zamrukk)
        self.axisY12.append(vakant1)
        self.axisY1.setLabelsFont(QFont("Times", 11))

        self.chart1.addAxis(self.axisY1, Qt.AlignLeft)
        self.series1.attachAxis(self.axisY1)

        self.set01.setColor(QColor("#397dd6"))
        self.set01.setLabelFont(QFont("Times", 11))

        self.set11.setColor(QColor("#b2ebf2"))
        self.set11.setLabelColor(QColor("#616161"))
        self.set11.setLabelFont(QFont("Times", 11))

        self.set21.setLabelColor(QColor("#616161"))
        self.set21.setLabelFont(QFont("Times", 11))

    def click_filtrotdel(self, int_xxx):
        self.series3.clear()
        self.axisY3.clear()
        self.axisY32.clear()
        dfcheckk3 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='O:S')
        df_checkk3 = dfcheckk3.dropna(subset=[dfcheckk3.columns[1]])
        filtered_df = df_checkk3.sort_values(by=df_checkk3.columns[int_xxx], ascending=False)

        self.set03 = QBarSet("Факт")
        self.set13 = QBarSet("Вакант")
        self.set23 = QBarSet("% Ваканта")
        zamrukk = []
        vakant3 = []
        for index, row in filtered_df[::-1].iterrows():
            label = row[0]
            zamrukk.append(label)
            value1 = int(row[1])
            value2 = int(row[3])
            okrugl = round((value2 / (value1 + value2) * 100), 1)
            value3 = str(okrugl) + str(index) + "%"
            self.set03.append(value1)
            self.set13.append(value2)
            vakant3.append(value3)

        self.series3.append(self.set03)
        self.series3.append(self.set13)
        # self.series3.append(self.set23)

        self.axisY3.append(zamrukk)
        self.axisY32.append(vakant3)
        self.axisY3.setLabelsFont(QFont("Times", 8))

        self.chart3.addAxis(self.axisY3, Qt.AlignLeft)
        self.series3.attachAxis(self.axisY3)

        self.set03.setColor(QColor("#397dd6"))
        self.set03.setLabelFont(QFont("Times", 8))

        self.set13.setColor(QColor("#b2ebf2"))
        self.set13.setLabelColor(QColor("#616161"))
        self.set13.setLabelFont(QFont("Times", 8))

        self.set23.setLabelColor(QColor("#616161"))
        self.set23.setLabelFont(QFont("Times", 8))

    def vosstanovit(self):
        self.series1.clear()
        self.axisY1.clear()
        self.axisY12.clear()

        dfcheckk = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='W:AA')
        df2_checkk = dfcheckk.dropna(subset=[dfcheckk.columns[1]])
        filtered_df = df2_checkk.sort_values(by=df2_checkk.columns[0], ascending=False)

        self.set01 = QBarSet("Факт")
        self.set11 = QBarSet("Вакант")
        self.set21 = QBarSet("% Ваканта")
        vakant1 = []
        zamrukk = []

        for index, row in filtered_df[::-1].iterrows():
            label = row[0]
            zamrukk.append(label)
            value1 = int(row[1])
            value2 = int(row[3])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            value3 = str(okrugl) + "%"
            self.set01.append(value1)
            self.set11.append(value2)
            vakant1.append(value3)

        self.series1.append(self.set01)
        self.series1.append(self.set11)

        self.axisY12.append(vakant1)
        self.axisY1.append(zamrukk)

        self.chart1.addAxis(self.axisY1, Qt.AlignLeft)
        self.series1.attachAxis(self.axisY1)

        self.chart1.addAxis(self.axisY12, Qt.AlignRight)
        self.series1.attachAxis(self.axisY12)

        self.set01.setColor(QColor("#397dd6"))
        self.set01.setLabelFont(QFont("Times", 11))

        self.set11.setColor(QColor("#b2ebf2"))
        self.set11.setLabelColor(QColor("#616161"))
        self.set11.setLabelFont(QFont("Times", 11))

        self.set21.setLabelColor(QColor("#616161"))
        self.set21.setLabelFont(QFont("Times", 11))

        # очищаем все фильтры которые ранее использовали
        self.series3.clear()
        self.axisY3.clear()
        self.axisY32.clear()

        dfcheckk3 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='N:S')
        df_checkk3 = dfcheckk3.dropna(subset=[dfcheckk3.columns[1]])

        self.set03 = QBarSet("Факт")
        self.set13 = QBarSet("Вакант")
        self.set23 = QBarSet("% Ваканта")
        zamrukk = []
        vakant3 = []
        for index, row in df_checkk3[::-1].iterrows():
            label = row[1]
            zamrukk.append(label)
            value1 = int(row[2])
            value2 = int(row[4])
            okrugl = round((value2 / (value1 + value2) * 100), 1)
            value3 = str(okrugl) + str(index) + "%"
            self.set03.append(value1)
            self.set13.append(value2)
            vakant3.append(value3)

        self.series3.append(self.set03)
        self.series3.append(self.set13)

        self.axisY3.append(zamrukk)
        self.axisY32.append(vakant3)

        self.axisY3.setLabelsFont(QFont("Times", 8))

        self.chart3.addAxis(self.axisY3, Qt.AlignLeft)
        self.series3.attachAxis(self.axisY3)

        self.set03.setColor(QColor("#397dd6"))
        self.set03.setLabelFont(QFont("Times", 11))

        self.set13.setColor(QColor("#b2ebf2"))
        self.set13.setLabelColor(QColor("#616161"))
        self.set13.setLabelFont(QFont("Times", 11))

        self.set23.setLabelColor(QColor("#616161"))
        self.set23.setLabelFont(QFont("Times", 11))

        self.series2.clear()
        self.axisY2.clear()
        self.axisY22.clear()
        self.chart2.setTitle("")

        dfcheckk2 = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='A:G')
        df2_checkk2 = dfcheckk2.dropna(subset=[dfcheckk2.columns[1]])
        grouped_df = df2_checkk2.groupby(by=df2_checkk2.columns[6]).sum()

        self.set02 = QBarSet("Факт")
        self.set12 = QBarSet("Вакант")
        self.set22 = QBarSet("% Ваканта")
        vakant2 = []
        zamrukk = []
        kolich = 0
        for index, row in grouped_df[::-1].iterrows():
            zamrukk.append(index)
            value1 = int(row[3])
            value2 = int(row[5])
            okrugl = round((value2 / (value1 + value2) * 100), 2)
            kolich += 1
            value3 = str(okrugl) + str(kolich) + "%"
            self.set02.append(value1)
            self.set12.append(value2)
            vakant2.append(value3)

        self.series2.append(self.set02)
        self.series2.append(self.set12)

        self.axisY2.append(zamrukk)
        self.axisY22.append(vakant2)

        self.series2.attachAxis(self.axisY2)

        self.chart2.addAxis(self.axisY22, Qt.AlignRight)
        self.series2.attachAxis(self.axisY22)

        self.series2.attachAxis(self.axisX2)

        self.set02.setColor(QColor("#397dd6"))
        self.set02.setLabelFont(QFont("Times", 11))

        self.set12.setColor(QColor("#b2ebf2"))
        self.set12.setLabelColor(QColor("#616161"))
        self.set12.setLabelFont(QFont("Times", 11))

        self.set22.setLabelColor(QColor("#616161"))
        self.set22.setLabelFont(QFont("Times", 11))

        self.chart3.setTitle("")

    def renewdash(self):
        # обновляем дашборд кароч сначала открыаем штатку там по формулам отрисована конкретная табличка ее копируем в файл с макросом и запусаем в нем макрос
        try:
            dfkopyshtatka = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Дашборд', skiprows=0, usecols='B:G')
            wb = xw.Book('addition/Дашборд.xlsm')
            sheet = wb.sheets('Штатка')
            sheet.range('A1').value = dfkopyshtatka
            wb.save('addition/Дашборд.xlsm')
            wb.close()
            excel = win32.Dispatch("Excel.Application")
            excel.Quit()

            vba_book = xw.Book('addition/Дашборд.xlsm')
            vba_macro2 = vba_book.macro("Module1.copyandsortcolumns2()")
            vba_macro2()
            vba_book.save()  # Сохраняем изменения в файле
            vba_book.close()  # Закрываем файл
            excel = win32.Dispatch("Excel.Application")
            excel.Quit()

            df_infos = pd.read_excel('addition/Дашборд.xlsm', sheet_name='Лист2', usecols='I:K')
            df_infos2 = df_infos.dropna(subset=[df_infos.columns[0]])
            datenowpokazat = df_infos2.iloc[3, 0]
            datenowpokazat = str(datenowpokazat)

            # создаем сообщение что подгрузили
            self.message_box3dash = QMessageBox(self)
            self.message_box3dash.setWindowTitle("Отчет об отработке")
            self.message_box3dash.setText(f"Дашборд обновлен данные обновлены на {datenowpokazat}")
            self.message_box3dash.setStandardButtons(QMessageBox.Ok)
            self.message_box3dash.show()

            # обновляем дату в подписи
            self.labeltimeobn.setText(f"Актуальность дашборда на\n {datenowpokazat}")
            self.vosstanovit()

            # закрыли эксельку
            excel = win32.Dispatch("Excel.Application")
            excel.Quit()

            excel = win32.Dispatch("Excel.Application")
            excel.Quit()
        except Exception as e:
            log_print(e, 'Дашборд упал')
            msg = QMessageBox(self)
            msg.setWindowTitle("Ошибка")
            msg.setText('Не удалось обновить дашборд, перезагрузите приложение и попробуйте еще раз.')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.show()



stylesheet976 = """
    Login_page_app {
        background-image: url(картинки/fon2.png); 
        background-repeat: no-repeat; 
        background-position: center;
    }
"""

if __name__ == '__main__':
    log_print('Приложение запущено')


    if '-t' in sys.argv:
        print('Test open')
        app = QApplication(sys.argv)
        w = Main_window()
        w.show()
        sys.exit(app.exec_())
    else:
        app = QApplication(sys.argv)
        app.setStyleSheet(stylesheet976)
        w = Login_page_app()
        w.show()
        # ex = App1()
        sys.exit(app.exec_())
    log_print('Приложение остановлено')
