import asyncio
import shutil
import sys
from datetime import datetime
from typing import List

import requests
import xlwings as xw
import numpy as np
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QBoxLayout, QVBoxLayout, QCheckBox, QApplication, QTableWidget, QTableWidgetItem, \
    QTableView, QGridLayout, QPushButton, QHBoxLayout, QDialog, QMessageBox
from PyQt5 import QtCore, QtWidgets
import win32com.client as win32
from openpyxl.reader.excel import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import json

from src.utils import log_print
from src.order import order

class Accept_order_tab(QWidget):
    '''
    Окно с принятием приказа.
    __init__        Забираем датафрейм, формируем лист чекбоксов. добавляем лайауты, соединяем их.
                    Запускаем отрисовку кнопок и таблицы.
    self.order_list_information     list из лист['приказы в работе']

    _draw_table         По датафрейму order_list... отрисовываем таблицу. Создаем новые чекбоксы, записываем их
                        в лист и отображаем.

    _draw_buttons       3 кнопки связываем их и на виджет

    _resize_event       Делаем ширину колонок и строк в табличке

    _download_orders    Загружает все данные в класс order. return: List[order]

    is_checkbox         Проверяет стоит ли галочка в чекбоксе

    update_orders       Запускает _download_orders -> _draw_table -> _resize_event. См. выше

    clear_order_sheet   удаляет из эксельки те приказы, которые отмечены. return: List[order]. Отмеченных приказов

    dialog_box          диалоговое окно с произвольным текстом

    message_nach        из файла temp2 достает все данные и отправляет по ним сообщения начальнику.

    message_all_changes Отправляет в конкретные отделы данные о переводе сотрудников во всех отделах 1-м сообщением.

    move_person         Двигает человека в штатке. Удаляет 1 из 1й столбца. Проставляет дату.

    fire_person         Удаляет человека из штатки и удаляет дату.

    sort_dataframe      хз зачем нужен

    commit_changes_database     закидывает в эксельку измененных людей. move_person | fire_person по приказу.

    clear_potential_close       Убирает 1 из первого столбца. WARNING: похоже это не нужно. Так как мы делаем это уже в
                                move & fire.

    save_staff_sheet            Перезаписывает штатку и сохраняет ее. Прокидывает ее в облако

    send_data                   Отправляет данные с изменением человека на сервер. WARNING: вроде работает.

    accept_orders               При нажатии кнопки принять приказ. clear_order_sheet -> Отправляет в temp2 ->
                                Запуск сообщений. Если не получилось возвращаем приказы с галками назад. Иначе ->
                                Отправляем данные на сервер. send_data -> commit_changes_database(orders_in_work) ->
                                download_orders -> draw_table.

    decline_orders              При отказе приказа. clear_potential_close -> save_staff_sheet -> _download_orders ->
                                _draw_table.


    Общее: Кажется я немного перемудрил с классом order и с некоторыми функциями. Работает страшно, но работает. =)
    '''
    def __init__(self, parent):
        """
        order_information: List[order]: Список всех приказов в работе.
        checkbox_list: List[int]: список нажатых чекбоксов.
        main_layout: Главный слой
        button_layout: Слой сверху с кнопками
        table_widget_layout: Слой со всей таблицей
        """
        super().__init__(parent)

        self.table_labels = ['Подразделение', 'Должность', 'ФИО', 'Статус', 'Новый отдел',
                             'Дата', 'Номер приказа', 'Утверждение']
        self._checkbox_list = []  # Все объекты чекбокса

        self.orders_list_information = self._download_orders()

        self.main_layout_order = QVBoxLayout(self)
        self.button_layout_order = QHBoxLayout(self)
        self.table_widget_orders = QTableWidget(self)


        self.main_layout_order.addLayout(self.button_layout_order)
        self.main_layout_order.addWidget(self.table_widget_orders)

        self._draw_buttons()
        self._draw_table()

        self._resize_event()
        self.setStyleSheet("""
                       QCheckBox{
                            margin-left: auto;
                       }
                       QTableWidget::item{
                            color: black;
                       }      
                        """)
        log_print('Инициализирована вкладка Приказы в работе')

    def _draw_table(self):

        self.table_widget_orders.setColumnCount(len(self.table_labels))
        self.table_widget_orders.setRowCount(len(self.orders_list_information))
        self.table_widget_orders.setHorizontalHeaderLabels(self.table_labels)


        self._checkbox_list = []

        for row in range(len(self.orders_list_information)):
            self.table_widget_orders.setItem(row, 0, QTableWidgetItem(self.orders_list_information[row].department))

            self.table_widget_orders.setItem(row, 1, QTableWidgetItem(self.orders_list_information[row].post))
            self.table_widget_orders.setItem(row, 2, QTableWidgetItem(self.orders_list_information[row].name))
            self.table_widget_orders.setItem(row, 3, QTableWidgetItem(self.orders_list_information[row].status))
            self.table_widget_orders.setItem(row, 4, QTableWidgetItem(self.orders_list_information[row].new_department))
            self.table_widget_orders.setItem(row, 5, QTableWidgetItem(self.orders_list_information[row].data))
            self.table_widget_orders.setItem(row, 6, QTableWidgetItem(self.orders_list_information[row].order_number))

            checkbox = QCheckBox()
            self._checkbox_list.append(checkbox)
            self.table_widget_orders.setCellWidget(row, 7, checkbox)
        for row in range(len(self.orders_list_information)):
            for col in range(len(self.table_labels)-1):
                self.table_widget_orders.item(row, col).setFlags(Qt.ItemIsEditable)


        self._resize_event()

    def _draw_buttons(self):
        self.button_order_accept = QPushButton('Утвердить', self)
        self.button_order_decline = QPushButton('Отказать', self)
        self.button_order_update = QPushButton('Обновить', self)

        self.button_order_accept.clicked.connect(self.accept_orders)
        self.button_order_decline.clicked.connect(self.decline_orders)
        self.button_order_update.clicked.connect(self.update_orders)

        self.button_layout_order.addWidget(self.button_order_accept, alignment=Qt.AlignmentFlag.AlignCenter)
        self.button_layout_order.addWidget(self.button_order_decline, alignment=Qt.AlignmentFlag.AlignCenter)
        self.button_layout_order.addWidget(self.button_order_update, alignment=Qt.AlignmentFlag.AlignCenter)

    def _resize_event(self):
        self.table_widget_orders.setWordWrap(True)
        self.table_widget_orders.resizeColumnToContents(True)
        for i in range(len(self.table_labels)-1):
            self.table_widget_orders.setColumnWidth(i, 220)
        self.table_widget_orders.resizeRowsToContents()

    def _download_orders(self):
        """
        добавляет всю новую информацию о действующих приказах в orders_information
        Вызывается, когда добавляется новый приказ

        Для того, чтобы отображать приказы

        :return: List[order]: загружает список всех приказов в работе
        """
        orders_list = []

        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Приказы в работе']
        # ws.cell(row=1, column=2).value
        data = ws.values
        next(data)
        for data_row in list(data):
            dep, post, name, os, np, d, on, dt, sd, n0 = data_row
            orders_list.append(order(dep, post, name, os, np, d, on, dt, sd))


        return orders_list

    def is_checkbox(self, checkbox: QCheckBox):
        """
            Находит все проставленные галочки
        """
        if checkbox.checkState():
            return True
        return False

    def update_orders(self):
        self.orders_list_information = self._download_orders()
        self._draw_table()
        self._resize_event()

    def clear_order_sheet(self):
        """
        :return: Возвращает ордеры, которые мы отметили
        """

        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Приказы в работе']
        counter = 0
        orders_result = []
        for i in range(len(self._checkbox_list)):
            if self.is_checkbox(self._checkbox_list[i]):
                orders_result.append(self.orders_list_information[i])
                a = i - counter + 2
                ws.delete_rows(a)
                counter += 1

        book.save('addition/kadrifile.xlsx')
        book.close()
        return orders_result

    def dialog_box(self, text):
        msg = QMessageBox(self)
        msg.setWindowTitle("Выполнено")
        msg.setText(text)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.show()

    def message_nach(self):
        outlook = win32.Dispatch('outlook.application')
        df = pd.read_excel('addition/temp2.xlsx')
        df_mail = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Историярасслылок2',usecols="B:E", skiprows=3)
        df.fillna('', inplace=True)
        df_mail = df_mail.rename(columns={'Структурное подразделение': 'Подразделение'})

        for index, row in df.iterrows():
            if row['Подразделение'] == '':
                name = df_mail[df_mail['Подразделение'] == row['Новый отдел']]['Начальник отдела/Первый по должности если он отсутствует'].to_string(index=False)
                email_adress = df_mail[df_mail['Подразделение'] == row['Новый отдел']]['Почта'].to_string(index=False)

                mail = outlook.CreateItem(0)
                mail.To = email_adress
                mail.Subject = 'Изменение штата отдела'
                mail.Body = f'Уважаемый (-ая) {name}. Изменения в {row["Новый отдел"]}\n'
                mail.Body += f'{row["ФИО"]}\n' \
                             f'Предыдущая должность: {row["Должность"]}\n' \
                             f'Нынешняя должность (статус): {row["Статус"]}\n' \
                             f'Новый отдел: {row["Новый отдел"]}\n' \
                             f'Дата: {row["Дата"]} \t Номер приказа: {row["Номер приказа"]}'
                mail.Send()

            elif row['Подразделение'] == row['Новый отдел'] or row['Новый отдел'] == '':
                name = df_mail[df_mail['Подразделение'] == row['Подразделение']]['Начальник отдела/Первый по должности если он отсутствует'].to_string(index=False)
                email_adress = df_mail[df_mail['Подразделение'] == row['Подразделение']]['Почта'].to_string(index=False)

                mail = outlook.CreateItem(0)
                mail.To = email_adress
                mail.Subject = 'Изменение штата отдела'
                mail.Body = f'Уважаемый (-ая) {name}. Изменения в {row["Подразделение"]}\n'
                mail.Body += f'{row["ФИО"]}\n' \
                             f'Предыдущая должность: {row["Должность"]}\n' \
                             f'Нынешняя должность (статус): {row["Статус"]}\n' \
                             f'Новый отдел: {row["Новый отдел"]}\n' \
                             f'Дата: {row["Дата"]} \t Номер приказа: {row["Номер приказа"]}'
                mail.Send()

            else:
                name_1o = df_mail[df_mail['Подразделение'] == row['Подразделение']]['Начальник отдела/Первый по должности если он отсутствует'].to_string(index=False)
                email_adress_1o = df_mail[df_mail['Подразделение'] == row['Подразделение']]['Почта'].to_string(index=False)

                name_2o = df_mail[df_mail['Подразделение'] == row['Новый отдел']]['Начальник отдела/Первый по должности если он отсутствует'].to_string(index=False)
                email_adress_2o = df_mail[df_mail['Подразделение'] == row['Новый отдел']]['Почта'].to_string(index=False)

                mail = outlook.CreateItem(0)
                mail.To = email_adress_1o
                mail.Subject = 'Изменение штата отдела'
                mail.Body = f'Уважаемый (-ая) {name_1o}. Изменения в {row["Подразделение"]}\n'
                mail.Body += f'{row["ФИО"]}\n' \
                             f'Предыдущая должность: {row["Должность"]}\n' \
                             f'Нынешняя должность (статус): {row["Статус"]}\n' \
                             f'Новый отдел: {row["Новый отдел"]}\n' \
                             f'Дата: {row["Дата"]} \t Номер приказа: {row["Номер приказа"]}'
                mail.Send()

                mail = outlook.CreateItem(0)
                mail.To = email_adress_2o
                mail.Subject = 'Изменение штата отдела'
                mail.Body = f'Уважаемый (-ая) {name_2o}. Изменения в {row["Новый отдел"]}\n'
                mail.Body += f'{row["ФИО"]}\n' \
                             f'Предыдущая должность: {row["Должность"]}\n' \
                             f'Нынешняя должность (статус): {row["Статус"]}\n' \
                             f'Новый отдел: {row["Новый отдел"]}\n' \
                             f'Дата: {row["Дата"]} \t Номер приказа: {row["Номер приказа"]}'
                mail.Send()

    def message_all_changes(self):
        outlook = win32.Dispatch('outlook.application')
        df = pd.read_excel('addition/temp2.xlsx')
        df_mail = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Историярасслылок2', usecols="B:E", skiprows=3)
        df.fillna('', inplace=True)
        df_mail = df_mail.rename(columns={'Структурное подразделение': 'Подразделение'})
        df_office = df_mail[(df_mail['Подразделение'] == 'Административный отдел')
                                | (df_mail['Подразделение'] == 'Специалист по вопросам мобилизационной подготовки') \
                                | (df_mail['Подразделение'] == 'Специалист по вопросам гражданской обороны') \
                                | (df_mail['Подразделение'] == 'Отдел внутреннего контроля и аудита') \
                                | (df_mail['Подразделение'] == 'Отдел режима секретности и безопасности информации') \
                                | (df_mail['Подразделение'] == 'Отдел сопровождения пользователей')]

        for index, row in df_office.iterrows():
            name_head_office = row['Начальник отдела/Первый по должности если он отсутствует']
            email_head_office = row['Почта']

            mail = outlook.CreateItem(0)
            mail.To = email_head_office
            mail.Subject = 'Изменение штата отдела'
            mail.Body = f"Уважаемый (-ая) {name_head_office}\n" \
                        f"Изменения штата одела:\n\n"
            for index, person in df.iterrows():
                name = person['ФИО']
                post = person['Должность']
                status = person['Статус']
                department = person['Подразделение']
                name_head_department = df_mail.loc[df_mail['Подразделение'] == person['Подразделение'],
                ['Начальник отдела/Первый по должности если он отсутствует']].to_string(index=False, header=False)
                if department == '':
                    department = '-'
                    name_head_department = '-'

                new_department = person['Новый отдел']
                name_head_new_department = df_mail.loc[df_mail['Подразделение'] == person['Новый отдел'],
                                                   ['Начальник отдела/Первый по должности если он отсутствует']].to_string(index=False, header=False)
                if new_department == '':
                    new_department = '-'
                    name_head_new_department = '-'
                data = person['Дата']
                order = person['Номер приказа']


                mail.Body += f"{name}.\nПредыдущая должность: {post} в {department}. Начальник отдела: {name_head_department}.\n" \
                             f"Текущая должость/статус: {status} в {new_department}. Начальник отдела: {name_head_new_department}.\n" \
                             f"Дата: {data}\t Номер приказа: {order}\n\n"
            mail.Send()

    def move_person(self, df_staff: pd.DataFrame, order_elem):

        match_new_post = df_staff[(df_staff['Подразделение'] == order_elem.new_department)
                                  & (df_staff['Должность'] == order_elem.status)
                                  & (df_staff['Unnamed: 0'] == 1)].index[0]

        match_prev_post = df_staff[(df_staff['Подразделение'] == order_elem.department)
                                   & (df_staff['Должность'] == order_elem.post)
                                   & (df_staff['ФИО'] == order_elem.name)
                                   ]

        df_staff.loc[match_new_post, 'ФИО'] = order_elem.name
        df_staff.loc[match_new_post, 'Unnamed: 0'] = np.nan
        df_staff.loc[match_new_post, 'Дата приказа'] = order_elem.data

        df_staff.loc[match_prev_post.index, 'ФИО'] = np.nan
        df_staff.loc[match_prev_post.index, 'Дата приказа'] = np.nan

    def fire_person(self, df_staff: pd.DataFrame, order_elem):
        match_post = df_staff[(df_staff['Подразделение'] == order_elem.department)
                              & (df_staff['Должность'] == order_elem.post)
                              & (df_staff['ФИО'] == order_elem.name)
                              ]
        df_staff.loc[match_post.index, 'ФИО'] = np.nan
        df_staff.loc[match_post.index, 'Дата приказа'] = np.nan

    def sort_dataframe(self, df_staff: pd.DataFrame, changed_posts: set):

        for dep, post in changed_posts:
            if (post != "Уволен") and (dep != None) and (post != None):
                first_row = df_staff[(df_staff['Подразделение'] == dep) & (df_staff['Должность'] == post)].index[0]
                last_row = df_staff[(df_staff['Подразделение'] == dep) & (df_staff['Должность'] == post)].index[-1]
                df_copy = df_staff[first_row:last_row + 1].copy()
                df_copy.sort_values(by='ФИО', inplace=True)
                df_staff = pd.concat([df_staff.iloc[:first_row], df_copy, df_staff.iloc[last_row + 1:]], axis=0)
            else:
                pass

        return df_staff

    def commit_changes_database(self, orders_in_work: List[order]):
        """
            Производит изменения базы данных, а потом их сохраняет
        """
        df_staff = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:E')
        changed_posts = set()
        for order_elem in orders_in_work:
            changed_posts.add((order_elem.department, order_elem.post))
            changed_posts.add((order_elem.new_department, order_elem.status))
            if order_elem.status != 'Уволен' and order_elem.status != 'Назначение':
                self.move_person(df_staff, order_elem)

            elif order_elem.status == 'Уволен':
                self.fire_person(df_staff, order_elem)

        df_staff = self.sort_dataframe(df_staff, changed_posts)
        self.save_staff_sheet(df_staff)

    def clear_potential_close(self, df_staff: pd.DataFrame, orders_in_work):
        for order_elem in orders_in_work:
            if order_elem.status != 'Уволен':
                match_new_post = df_staff[(df_staff['Подразделение'] == order_elem.new_department)
                                          & (df_staff['Должность'] == order_elem.status)
                                          & (df_staff['Unnamed: 0'] == 1)].index[-1]

                df_staff.loc[match_new_post, 'Unnamed: 0'] = np.nan

    def save_staff_sheet(self, df_staff: pd.DataFrame):
        """Сохраняет книгу, открывает эксель сохраняет ее, кидает это все в облако"""

        book = load_workbook('addition/kadrifile.xlsx')
        ws = book['Штатка']
        ws.delete_cols(1, 8)
        ws.delete_rows(1, 500)
        for r in dataframe_to_rows(df_staff, index=False, header=True):
            ws.append(r)

        book.save('addition/kadrifile.xlsx')
        book.save(f"логи/File XYZ_{datetime.now().strftime('%Y-%m-%d %H.%M.%S')}.xlsx")
        book.close()

        try:
            with open('addition/путь.txt', 'r', encoding="utf-8") as file:
                put = file.readline().strip('\n')
                src_file = 'addition/temp.xlsx'
                dst_folder = f'{put}'
                shutil.copy(src_file, dst_folder)
                src_file = 'addition/kadrifile.xlsx'
                dst_folder = f'{put}'
                shutil.copy(src_file, dst_folder)
        except Exception as e:
            log_print(f'Не найдена облачная папка. {e}')
            self.dialog_box('Неполучилось найти папку в облаке')

    def send_data(self, orders: List):
        posts = requests.get("http://192.168.130.89:8000/staff/post")
        deps = requests.get("http://192.168.130.89:8000/staff/dep")
        posts = json.loads(posts.text)
        deps = json.loads(deps.text)

        status_mapping = {v: {'action': 1, 'is_active': True, 'post_id': k} for k, v in posts.items()}
        department_mapping = {v:k for k,v in deps.items()}

        data = {'users': []}
        i = 0
        for order in orders:
            status_info = status_mapping.get(order.status, {})
            depart_info = int(department_mapping.get(order.new_department, {}))
            is_active = status_info.get('is_active')

            if order.department == None:
                action = 3
            elif order.status == 'Уволен':
                action = 2
                is_active = False
            else:
                action = 1
            post_id = status_info.get('post_id')
            department_id = depart_info

            ord_json = {
                "action": action,
                "user_name": order.name,
                "post_id": post_id,
                "department_id": department_id,
                'is_active': is_active
            }


            data['users'].append(ord_json) # TODO: Кривой цикл, нужно переделать сервер

            i += 1
        ans = requests.post("http://192.168.130.89:8000/staff/", json=data)
        log_print(ans.text)


    def accept_orders(self):
        """
        Исполняет все выбранные приказы. Проходит по галочкам
        и исполняет код с ними.
        """
        orders_in_work = self.clear_order_sheet()
        df_to_temp = pd.DataFrame(columns=['Unnamed: 0', 'Подразделение',
                                           'Должность', 'ФИО', 'Статус', 'Новый отдел',
                                           'Дата', 'Номер приказа', 'Тип документа',
                                           'Дата заявления'])

        for i in range(len(orders_in_work)):
            ord = orders_in_work[i].get_order_info_to_pandas()
            ord.insert(0, '0')
            df_to_temp.loc[df_to_temp.shape[0]] = ord

        df_to_temp.to_excel("addition/temp2.xlsx", index=False)
        app = xw.App(visible=False)
        wb = xw.Book('addition/kadrifile.xlsx')
        wb.save('addition/kadrifile.xlsx')
        wb.close()
        app.quit()
        try:
            pass
            self.message_nach()
            self.message_all_changes()
        except Exception as e:
            book = load_workbook('addition/kadrifile.xlsx')
            ws = book['Приказы в работе']

            fastdf = orders_in_work
            for r in fastdf:
                ws.append(r.get_order_info_to_pandas())
            book.save('addition/kadrifile.xlsx')
            book.close()

            self.dialog_box(f'{e} | Сообщения не были отправлены')
            log_print(e, 'message error')
        else:
            try:
                self.send_data(orders_in_work)
            except Exception as e:
                log_print('Код с сервером не сработал', e)

            df_to_temp = pd.read_excel('addition/temp.xlsx')
            for i in range(len(orders_in_work)):
                ord = orders_in_work[i].get_order_info_to_pandas()
                ord.insert(0, '0')
                df_to_temp.loc[df_to_temp.shape[0]] = ord
            df_to_temp.to_excel("addition/temp.xlsx", index=False)

            self.commit_changes_database(orders_in_work)
            self.dialog_box('Приказы утверждены')

            self.orders_list_information = self._download_orders()
            self._draw_table()

    def decline_orders(self):
        df_staff = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
        orders_in_work = self.clear_order_sheet()
        self.clear_potential_close(df_staff, orders_in_work)
        self.save_staff_sheet(df_staff)
        self.dialog_box('Приказы отменены')
        self.orders_list_information = self._download_orders()
        self._draw_table()

