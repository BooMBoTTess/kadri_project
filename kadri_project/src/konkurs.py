import os
import sys
from typing import Dict, List
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
import random
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import shutil
import pandas as pd
import numpy as np
import xlwings as xw
import win32com.client as win32
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QColor, QPixmap
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QScrollArea
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout
import sys
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtChart import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPainter
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtChart import (QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis)

from src.utils import log_print

'''
download функции наверно можно использовать везде и в других прогах. Взять этот код и вывести его в отдельный py файл
и импортить везде.
'''

def download_globalTolocal(filename: str):
    with open('addition/путь.txt', 'r', encoding="utf-8") as file:
        put = file.readline().rstrip()
    # Путь к первому документу
    src_file = f'{put}/{filename}.xlsx'
    # Путь ко второму документу
    dst_file = f"addition/{filename}.xlsx"

    shutil.copy(src_file, dst_file)


def download_localToglobal(filename: str):
    with open('addition/путь.txt', 'r', encoding="utf-8") as file:
        put = file.readline().rstrip()
    # Путь к первому документу
    dst_file = f'{put}/{filename}.xlsx'
    # Путь ко второму документу
    src_file = f"addition/{filename}.xlsx"

    shutil.copy(src_file, dst_file)


class KonkursTabulation(QTabWidget):
    """
    Табуляция конкурса

    __init__            Создаем 3 табы, добавляем. Скачиваем конкурс с диска X. Запускаем функции с отображением
                        графического интерфейса. 3х таб

    tabWorkers_listUI   Получает датафрейм (хз какой). Отображает окно с формированием работников /штатки

    tabKomicia_listUI   Отображает окно с формированием комиссии

    tabcurrent_contestUI            Отображает окно с текущим конкурсом.
    df_contest                      Датафрейм с текущим конкурсом.

    fill_contest_TableWidget        Заполняет tableWidget с конкурсом. По датафрейму, который передается

    preproc_df_contest              Делает все вычисления по конкурсу и формирует df в tablewidget. return pd.Dataframe


    """

    def __init__(self, df, parent=None):
        super(KonkursTabulation, self).__init__(parent)
        download_globalTolocal('конкурс')
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tab3 = QWidget()
        self.addTab(self.tab1, "Формирование конкурса (Работники)")
        self.addTab(self.tab2, "Формирование конкурса (Комиссия)")
        self.addTab(self.tab3, "Формирование конкурса (Комиссия)")
        self.tabWorkers_listUI(df)
        self.tabKomicia_listUI(df)
        self.tabcurrent_contestUI()

    def tabWorkers_listUI(self, df):
        self.layout100 = QVBoxLayout()
        self.h_layout2 = QHBoxLayout()
        self.h_layout = QHBoxLayout()
        self.h_layout4 = QHBoxLayout()
        self.df = df
        self.table_widget23 = QTableWidget(self)
        self.table_widget23.setEditTriggers(QTableWidget.AllEditTriggers)

        self.table_widget9999 = QTableWidget(self)
        self.table_widget9999.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget23.setStyleSheet("background-image: url(картинки/2.png)")
        self.table_widget9999.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget23.setFont(font)
        self.table_widget9999.setFont(font)
        self.table_widget23.setFixedHeight(420)
        self.table_widget9999.setFixedHeight(420)

        self.label1 = QLabel("Поиск:")
        self.label1.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                  "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                  "font-weight: bold")
        self.label1.setFixedHeight(20)

        self.label33 = QLabel("Перечень сотрудников добавленных в конкурс:")
        self.label33.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                   "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                   "font-weight: bold")
        self.label33.setFixedHeight(27)

        self.searchLineEdit = QLineEdit()
        self.searchLineEdit.textChanged.connect(self.search_tablerabot)
        self.label1.setFixedHeight(20)

        self.button_addcheliki = QPushButton("Добавить сотрудника в конкурс")
        self.button_addcheliki.clicked.connect(self.add_chelikirabot)
        self.button_addcheliki.setFixedHeight(35)
        self.button_addcheliki.setObjectName('blue_button')

        self.button_deletecheliki = QPushButton("Удалить сотрудника из конкурса")
        self.button_deletecheliki.clicked.connect(self.delete_chelikirabot)
        self.button_deletecheliki.setFixedHeight(35)
        self.button_deletecheliki.setObjectName('blue_button')

        self.button_newcheliki = QPushButton("Добавить Нового Сотрудника в конкурс")
        self.button_newcheliki.clicked.connect(self.new_chelikirabot)
        self.button_newcheliki.setFixedHeight(35)
        self.button_newcheliki.setObjectName('blue_button')

        self.h_layout.addWidget(self.button_addcheliki)
        self.h_layout.addWidget(self.button_newcheliki)
        self.h_layout.addWidget(self.button_deletecheliki)

        self.layout100.addLayout(self.h_layout2)
        self.layout100.addLayout(self.h_layout)
        self.layout100.addLayout(self.h_layout4)

        self.h_layout4.addWidget(self.label1)
        self.h_layout4.addWidget(self.searchLineEdit)
        self.layout100.addWidget(self.table_widget23)
        self.layout100.addWidget(self.label33)
        self.layout100.addWidget(self.table_widget9999)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelikrabot()

        self.tab1.setLayout(self.layout100)

    def tabKomicia_listUI(self, df):
        self.layout200 = QVBoxLayout()
        self.h_layout5 = QHBoxLayout()
        self.h_layout6 = QHBoxLayout()
        self.h_layout7 = QHBoxLayout()
        self.h_layout8 = QHBoxLayout()
        self.df = df
        self.table_widget33 = QTableWidget(self)
        self.table_widget33.setEditTriggers(QTableWidget.AllEditTriggers)

        self.table_widget99 = QTableWidget(self)
        self.table_widget99.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table_widget33.setStyleSheet("background-image: url(картинки/2.png)")
        self.table_widget99.setStyleSheet("background-image: url(картинки/2.png)")
        font = QFont("Times", 10)
        self.table_widget33.setFont(font)
        self.table_widget99.setFont(font)
        self.table_widget33.setFixedHeight(400)
        self.table_widget99.setFixedHeight(400)

        self.label111 = QLabel("Поиск:")
        self.label111.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                    "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                    "font-weight: bold")
        self.label111.setFixedHeight(20)

        self.label333 = QLabel("Перечень членов комиссии добавленных в конкурс:")
        self.label333.setStyleSheet("\n" "color: rgb(33, 53, 89);\n"
                                    "\n" "background-color: rgb(255, 255, 255,0);\n" "\n" "font: 16pt \"Open Sans Light\";\n" "\n"
                                    "font-weight: bold")
        self.label333.setFixedHeight(30)

        self.searchLineEdit1 = QLineEdit()
        self.searchLineEdit1.textChanged.connect(self.search_tablekomic)

        self.button_addcheliki1 = QPushButton("Добавить члена комиссии в конкурс")
        self.button_addcheliki1.clicked.connect(self.add_chelikikomic)
        self.button_addcheliki1.setFixedHeight(35)
        self.button_addcheliki1.setObjectName('blue_button')

        self.button_deletecheliki1 = QPushButton("Удалить члена комиссии из конкурса")
        self.button_deletecheliki1.clicked.connect(self.delete_chelikikomic)
        self.button_deletecheliki1.setFixedHeight(35)
        self.button_deletecheliki1.setObjectName('blue_button')

        self.button_newcheliki1 = QPushButton("Добавить Нового члена комиссии в конкурс")
        self.button_newcheliki1.clicked.connect(self.new_chelikikomic)
        self.button_newcheliki1.setFixedHeight(35)
        self.button_newcheliki1.setObjectName('blue_button')

        self.button_izmencheliki1 = QPushButton("Начать работу над конкурсом")
        self.button_izmencheliki1.setFixedHeight(35)
        self.button_izmencheliki1.clicked.connect(self.izmenit_cheliki)

        self.h_layout6.addWidget(self.button_addcheliki1)
        self.h_layout6.addWidget(self.button_deletecheliki1)

        self.h_layout7.addWidget(self.button_newcheliki1)
        self.h_layout7.addWidget(self.button_izmencheliki1)

        self.layout200.addLayout(self.h_layout5)
        self.layout200.addLayout(self.h_layout6)
        self.layout200.addLayout(self.h_layout7)
        self.layout200.addLayout(self.h_layout8)

        self.h_layout8.addWidget(self.label111)
        self.h_layout8.addWidget(self.searchLineEdit1)
        self.layout200.addWidget(self.table_widget33)
        self.layout200.addWidget(self.label333)
        self.layout200.addWidget(self.table_widget99)

        self.dialogs = list()  # Список под окна класса WindowGRBS
        self.dialogs_last = list()  # Список под окна для открытие выбора из логов
        self.table_view = QTableView()
        self.zapoln_chelikkomic()

        self.tab2.setLayout(self.layout200)

    def tabcurrent_contestUI(self):
        df_contest = pd.read_excel('addition/конкурс.xlsx', sheet_name='Конкурс (Работники)')
        df_contest = self.preproc_df_contest(df_contest)
        self.main_layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.tab3.setLayout(self.main_layout)
        self.buttons_layout = QHBoxLayout()
        self.main_layout.addLayout(self.buttons_layout)

        self.table_widget.setFont(QFont("Times", 10))
        self.table_widget.setColumnCount(df_contest.shape[1])
        self.table_widget.setHorizontalHeaderLabels(df_contest.columns.to_list())
        self.fill_contest_TableWidget(self.table_widget, df_contest)
        self.main_layout.addWidget(self.table_widget)


        self.submit_button = QPushButton("Закончить текущий конкурс")
        self.submit_button.clicked.connect(self.onExcelButton_click)
        self.buttons_layout.addWidget(self.submit_button)

    def fill_contest_TableWidget(self, tw, df):
        self.table_widget.blockSignals(True)
        self.table_widget.setRowCount(len(df))
        for i, row in df.iterrows():
            for j in range(len(row)):
                elem = str(df.iloc[i, j])
                if elem == 'nan':
                    elem = ''
                cell = QTableWidgetItem(elem)
                cell.setFlags(Qt.ItemIsEnabled)
                self.table_widget.setItem(i, j, cell)
        self.table_widget.resizeColumnsToContents()
        self.table_widget.setColumnWidth(1, 200)
        self.table_widget.setColumnWidth(3, 400)
        self.table_widget.blockSignals(False)

    def preproc_df_contest(self, df_contest):
        # Комиссия
        df_contest['Баллы комиссия'] = df_contest.loc[:, df_contest.columns.str.endswith('Балл')].sum(axis=1) \
                                              / df_contest.loc[:, df_contest.columns.str.endswith('Балл')].count(axis=1)
        df_contest['Баллы комиссия'] = df_contest['Баллы комиссия'].round(2)
        df_contest['Итог решений комиссии'] = df_contest.loc[:,
                                              df_contest.columns.str.endswith('ЗаПротив')].sum(axis=1) \
                                              / df_contest.loc[:, df_contest.columns.str.endswith('ЗаПротив')] \
                                                  .count(axis=1)
        df_contest['Решение комиссии'] = np.where(df_contest['Итог решений комиссии'] > 0.5, 'За', 'Против')

        # Тестирование
        df_contest['% за тестирование'] = df_contest.loc[:, 'Количество ответов'] / df_contest.loc[:,
                                                                                    'Количество вопросов']
        df_contest['% за тестирование'] = df_contest['% за тестирование'].round(2) * 100


        # Пересобирание колонок датафрейма

        df_contest = df_contest.loc[:, ['ФИО', 'Подразделение', 'Должность', 'Новый отдел', 'Претендующая Должность',
                                        'Решение комиссии', '% за тестирование',  'Баллы комиссия']]

        return df_contest


    def izmenit_cheliki(self):
        try:
            nepr = 0
            rowData333 = 1
            for row in range(self.table_widget9999.rowCount()):
                rowData333 = 0
                for col in range(self.table_widget9999.columnCount()):
                    if col != 3 and col != 4 and col != 6 and col != 7:
                        if self.table_widget9999.item(row, col).text() == '':
                            nepr += 1
                    elif col == 3 or col == 4:
                        if self.table_widget9999.cellWidget(row, col).currentText() == '':
                            nepr += 1
            rowData3333 = 1
            for row in range(self.table_widget99.rowCount()):
                rowData3333 = 0
                for col in range(self.table_widget99.columnCount()):
                    if col != 3:
                        if self.table_widget99.item(row, col).text() == '':
                            nepr += 1
                    elif col == 3:
                        if self.table_widget99.cellWidget(row, col).currentText() == '':
                            nepr += 1
            if nepr == 0 and rowData333 == 0 and rowData3333 == 0:
                log_print('init contest')
                self.start_work_order()
            elif rowData3333 == 1 and rowData333 == 1:
                self.message_box9616 = QMessageBox()
                self.message_box9616.setWindowTitle("Отчет об ошибках")
                self.message_box9616.setText(
                    "Заполните все ПУСТЫЕ таблицы конкурса!")
                self.message_box9616.setStandardButtons(QMessageBox.Ok)
                self.message_box9616.show()
            elif rowData3333 == 0 and rowData333 == 1:
                self.message_box9616 = QMessageBox()
                self.message_box9616.setWindowTitle("Отчет об ошибках")
                self.message_box9616.setText(
                    "Заполните все ПУСТЫЕ в таблице конкурса Работники!")
                self.message_box9616.setStandardButtons(QMessageBox.Ok)
                self.message_box9616.show()
            elif rowData3333 == 1 and rowData333 == 0:
                self.message_box9616 = QMessageBox()
                self.message_box9616.setWindowTitle("Отчет об ошибках")
                self.message_box9616.setText(
                    "Заполните все ПУСТЫЕ в таблице конкурса Комиссия!")
                self.message_box9616.setStandardButtons(QMessageBox.Ok)
                self.message_box9616.show()
            else:
                self.message_box9616 = QMessageBox()
                self.message_box9616.setWindowTitle("Отчет об ошибках")
                self.message_box9616.setText(
                    "Заполните все ПУСТЫЕ поля согласно предоставленной инструкции!")
                self.message_box9616.setStandardButtons(QMessageBox.Ok)
                self.message_box9616.show()
        except:
            self.message_box9616 = QMessageBox()
            self.message_box9616.setWindowTitle("Отчет об ошибках")
            self.message_box9616.setText(
                "Заполните все ПУСТЫЕ поля согласно предоставленной инструкции!")
            self.message_box9616.setStandardButtons(QMessageBox.Ok)
            self.message_box9616.show()

    def get_values_from_BottomSheetrabot(self):
        '''Подгрузка из нижней таблицы данных'''

        columnnamesr = ['Подразделение', 'Должность', 'ФИО', 'Претендующая Должность', 'Новый отдел',
                        'Количество вопросов', 'Дата', 'Время', 'Логин', 'Пароль']
        df_bottom_sheetr = pd.DataFrame(columns=columnnamesr)
        for row in range(self.table_widget9999.rowCount()):
            rowData = []
            for col in range(self.table_widget9999.columnCount()):
                if col != 3 and col != 4 and col != 6 and col != 7:
                    if col == 2:
                        text_names = self.table_widget9999.item(row, col).text()
                    zzz = self.table_widget9999.item(row, col).text()
                    rowData.append(zzz)
                elif col == 3 or col == 4:
                    zzz2 = self.table_widget9999.cellWidget(row, col).currentText()
                    rowData.append(zzz2)
                elif col == 6 or col == 7:
                    zzz3 = self.table_widget9999.cellWidget(row, col).text()
                    rowData.append(zzz3)

            text_2 = text_names.split("\n")
            for i in text_2:
                text_3 = (i.split()[0])
                rowData.append(text_3)
                # password GENERATOR

                password = ''.join([chr(random.randint(97, 97 + 25)) for i in range(5)])
                rowData.append(password)

            df_bottom_sheetr = pd.concat([df_bottom_sheetr, pd.DataFrame([rowData], columns=columnnamesr)])

        return df_bottom_sheetr

    def get_values_from_BottomSheetkomic(self):
        '''Подгрузка из нижней таблицы данных'''

        columnnamesk = ['Подразделение', 'Должность', 'ФИО', 'Статус участника конкурсной комиссии', 'Логин', 'Пароль']
        df_bottom_sheetk = pd.DataFrame(columns=columnnamesk)
        for row in range(self.table_widget99.rowCount()):
            rowDatak = []
            for col in range(self.table_widget99.columnCount()):
                if col != 3:
                    if col == 2:
                        text_names1 = []
                        text_names1 = self.table_widget99.item(row, col).text()
                    zzz5 = self.table_widget99.item(row, col).text()
                    rowDatak.append(zzz5)
                elif col == 3:
                    zzz6 = self.table_widget99.cellWidget(row, col).currentText()
                    rowDatak.append(zzz6)

            text_4 = text_names1.split("\n")
            for i in text_4:
                text_5 = (i.split()[0])
                rowDatak.append(text_5)
            alphabet = '6432'
            password = '2314'
            rowDatak.append(password)

            df_bottom_sheetk = pd.concat([df_bottom_sheetk, pd.DataFrame([rowDatak], columns=columnnamesk)])

        return df_bottom_sheetk

    def start_work_order(self):
        """
        Создает конкурс и документ
        :return:
        """

        log_print('Начата работа над конкурсом.')

        try:
            df_bottom_sheetrabot = self.get_values_from_BottomSheetrabot()
            df_bottom_sheetkomic = self.get_values_from_BottomSheetkomic()
            self.zapoln_chelikrabot()
            self.zapoln_chelikkomic()
            self.message_box96 = QMessageBox()
            self.message_box96.setWindowTitle("Выполнено")
            self.message_box96.setText(
                "Таблица Конкурса Заполнена.")
            self.message_box96.setStandardButtons(QMessageBox.Ok)
            self.message_box96.show()

        except Exception as e:
            msg = QMessageBox(self)
            msg.setWindowTitle("Ошибка")
            msg.setText(f"{e}. Невозможно выполнить запрос.")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.show()
            log_print(e, '| Невозможно выполнить запрос.')
            print(e, '| Невозможно выполнить запрос.')

        df_bottom_sheetrabot.insert(6, 'Количество ответов', '')

        for elem in df_bottom_sheetkomic['ФИО']:
            df_bottom_sheetrabot.insert(11, elem + ' Мотивировка', '')
            df_bottom_sheetrabot.insert(11, elem + ' ЗаПротив', '')
            df_bottom_sheetrabot.insert(11, elem + ' Балл', '')

        wb = Workbook()
        ws = wb.active
        ws.title = "Конкурс (Работники)"
        wb.create_sheet("Конкурс (Комиссия)")
        ws = wb["Конкурс (Работники)"]

        rows = dataframe_to_rows(df_bottom_sheetrabot, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = wb["Конкурс (Комиссия)"]
        rows = dataframe_to_rows(df_bottom_sheetkomic, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save('addition/конкурс.xlsx')
        wb.close()
        # self.message_nach()
        download_localToglobal('конкурс')

        log_print(f'Работа над конкурсом закончилась. Успешно обработано {len(df_bottom_sheetrabot)} запросов.')

    def message_nach(self):
        outlook = win32.Dispatch('outlook.application')
        df_teste = pd.read_excel('addition/конкурс.xlsx', sheet_name='Конкурс (Работники)')
        df_mail = pd.read_excel('addition/конкурс_почта .xlsx', sheet_name='email')
        df_teste.fillna('', inplace=True)
        # df_mail = df_mail.rename(columns={'Структурное подразделение': 'Подразделение'})

        for index, row in df_teste.iterrows():
            if row['Подразделение'] != '':
                name = df_teste['ФИО'].to_string(index=False)
                email_adress = df_mail['Почта'].to_string(index=False)
                mail = outlook.CreateItem(0)
                mail.To = email_adress
                mail.Subject = 'Приглашение на тестирование'
                mail.Body = f'Уважаемый (-ая) {name}. Изменения в {row["Новый отдел"]}\n'
                mail.Body += f'Приглашаем вас на конкурс: \n' \
                             f'Предыдущая должность: {row["Должность"]}\n' \
                             f'Претендующая должность: {row["Претендующая Должность"]}\n' \
                             f'Новый отдел: {row["Новый отдел"]}\n' \
                             f'Дата: {row["Дата"]} \t Время: {row["Время"]} \t Логин: {row["Логин"]} \t Пароль: {row["Пароль"]}'
                mail.Send()

    def delete_chelikirabot(self):
        row = self.table_widget9999.currentRow()
        if row > -1:  # Если есть выделенная строка/элемент
            self.table_widget9999.removeRow(row)
            # Следующий вызов нужен для того, чтобы
            # сбросить индекс выбранной строки (чтобы currentRow установился в -
            log_print(f'строка {row} удалена из нижнего виджета')
            self.table_widget9999.selectionModel().clearCurrentIndex()

        else:
            current_row_count = self.table_widget9999.rowCount()
            self.table_widget9999.setRowCount(current_row_count - 1)
            log_print(f'строка{row} удалена из нижнего виджета')

    def delete_chelikikomic(self):
        row = self.table_widget99.currentRow()
        if row > -1:  # Если есть выделенная строка/элемент
            self.table_widget99.removeRow(row)
            # Следующий вызов нужен для того, чтобы
            # сбросить индекс выбранной строки (чтобы currentRow установился в -
            log_print(f'строка {row} удалена из нижнего виджета')
            self.table_widget99.selectionModel().clearCurrentIndex()
        else:
            current_row_count = self.table_widget99.rowCount()
            self.table_widget99.setRowCount(current_row_count - 1)
            log_print(f'строка{row} удалена из нижнего виджета')

    def add_chelikirabot(self):
        try:
            current_row_count = self.table_widget9999.rowCount()
            self.table_widget9999.setColumnCount(8)
            self.table_widget9999.setRowCount(current_row_count + 1)
            combo = QComboBox()
            combo.addItems(
                ["", "Должность 1", "Должность 2", "Должность 3"])
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)

            combo_2 = QComboBox()
            combo_2.setEditable(True)
            combo_2.setInsertPolicy(QComboBox().InsertAfterCurrent)
            combo_2.addItems(["", "Отдел 1", "Отдел 2", "Отдел 3"])
            self.table_widget9999.setCellWidget(current_row_count, 4, combo_2)

            selected_row = self.table_widget23.currentRow()
            row_items = []

            for column in range(self.table_widget23.columnCount()):
                item = self.table_widget23.item(selected_row, column)
                row_items.append(item.text())
            row_position = self.table_widget9999.rowCount()

            for column, item in enumerate(row_items):
                self.table_widget9999.setItem(row_position - 1, column, QTableWidgetItem(item))

            otdel_bled = row_items[0]
            self.table_widget9999.setItem(row_position - 1, 4, combo_2.setCurrentText(otdel_bled))

            self.table_widget9999.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
            self.table_widget9999.setHorizontalHeaderLabels(
                ["Подразделение", "Должность", "ФИО", "Претендующая Должность", "Новый отдел", "Количество вопросов",
                 'Дата', 'Время'])

            self.table_widget9999.setItem(current_row_count, 5, QTableWidgetItem('40'))
            calendarr = QCalendarWidget()
            calendarr.setGridVisible(True)
            calendarr.setStyleSheet("QCalendarWidget QToolButton"
                                    "{"
                                    "background-color : lightgrey;"
                                    "color : black"
                                    "}")

            calendarchik = QtWidgets.QDateEdit(calendarPopup=True)
            calendarchik.setCalendarWidget(calendarr)
            calendarchik.setDateTime(QtCore.QDateTime.currentDateTime())
            calendarchik.setDisplayFormat('dd.MM.yyyy')
            self.table_widget9999.setCellWidget(row_position - 1, 6, calendarchik)

            time = QtWidgets.QDateTimeEdit()
            time.setDisplayFormat('HH:mm')
            time.setTime(QTime(12, 0))
            self.table_widget9999.setCellWidget(row_position - 1, 7, time)

            self.table_widget9999.setWordWrap(True)
            self.table_widget9999.resizeColumnsToContents()
            self.table_widget9999.setColumnWidth(0, 180)
            self.table_widget9999.setColumnWidth(1, 90)
            self.table_widget9999.setColumnWidth(2, 180)
            self.table_widget9999.setColumnWidth(5, 180)
            self.table_widget9999.resizeRowsToContents()
            self.table_widget9999.setColumnWidth(7, 90)

            log_print('Добавлено поле для ввода приказа')
        except Exception as e:
            log_print(e, 'Ошибка при добавлении поля для ввода приказа')
            pass

    def add_chelikikomic(self):
        try:
            current_row_count = self.table_widget99.rowCount()
            self.table_widget99.setColumnCount(4)
            self.table_widget99.setRowCount(current_row_count + 1)
            combo = QComboBox()
            combo.addItems(
                ["", "Назначения", "Председатель конкурсной комиссии", "Заместитель председателя комиссии",
                 "Секретарь конкурсной комиссии", "Член конкурсной комиссии", "Независимый эксперт",
                 "Наблюдатель"])
            self.table_widget99.setCellWidget(current_row_count, 3, combo)
            self.table_widget99.setCellWidget(current_row_count, 3, combo)

            selected_row = self.table_widget33.currentRow()
            row_items = []

            for column in range(self.table_widget33.columnCount()):
                item = self.table_widget33.item(selected_row, column)
                row_items.append(item.text())
            row_position = self.table_widget99.rowCount()

            for column, item in enumerate(row_items):
                self.table_widget99.setItem(row_position - 1, column, QTableWidgetItem(item))

            self.table_widget99.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
            self.table_widget99.setHorizontalHeaderLabels(
                ["Подразделение", "Должность", "ФИО", "Статус участника конкурсной комиссии"])

            self.table_widget99.setWordWrap(True)
            self.table_widget99.resizeColumnsToContents()
            self.table_widget99.setColumnWidth(0, 180)
            self.table_widget99.setColumnWidth(1, 90)
            self.table_widget99.setColumnWidth(2, 180)
            self.table_widget99.setColumnWidth(5, 180)
            self.table_widget99.resizeRowsToContents()

            log_print('Добавлено поле для ввода приказа')
        except Exception as e:
            log_print(e, 'Ошибка при добавлении поля для ввода приказа')
            pass

    def new_chelikirabot(self):
        try:
            current_row_count = self.table_widget9999.rowCount()
            self.table_widget9999.setColumnCount(8)
            self.table_widget9999.setRowCount(current_row_count + 1)
            combo = QComboBox()
            combo.addItems(
                ["", "Назначения", "Начальник отдела", "Заместитель начальника отдела", "Консультант",
                 "Главный казначей",
                 "Главный специалист-эксперт",
                 "Старший казначей", "Ведущий специалист-эксперт", "Казначей", "Специалист 1 разряда"])
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)
            self.table_widget9999.setCellWidget(current_row_count, 3, combo)

            combo_2 = QComboBox()
            combo_2.setEditable(True)
            # combo_2.setWordWrap(True)
            combo_2.setInsertPolicy(QComboBox().InsertAfterCurrent)
            combo_2.addItems(["", "Отдел 1", "Отдел 2", "Отдел 3"])
            self.table_widget9999.setCellWidget(current_row_count, 4, combo_2)

            selected_row = self.table_widget23.currentRow()
            row_items = []

            for column in range(self.table_widget23.columnCount()):
                item = self.table_widget23.item(selected_row, column)
            row_position = self.table_widget9999.rowCount()

            self.table_widget9999.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
            self.table_widget9999.setHorizontalHeaderLabels(
                ["Подразделение", "Должность", "ФИО", "Претендующая Должность", "Новый отдел", "Количество вопросов",
                 'Дата', 'Время'])

            self.table_widget9999.setItem(current_row_count, 5, QTableWidgetItem('40'))

            calendarr = QCalendarWidget()
            calendarr.setGridVisible(True)
            calendarr.setStyleSheet("QCalendarWidget QToolButton"
                                    "{"
                                    "background-color : lightgrey;"
                                    "color : black"
                                    "}")

            calendarchik = QtWidgets.QDateEdit(calendarPopup=True)
            calendarchik.setCalendarWidget(calendarr)
            calendarchik.setDateTime(QtCore.QDateTime.currentDateTime())
            calendarchik.setDisplayFormat('dd.MM.yyyy')
            self.table_widget9999.setCellWidget(row_position - 1, 6, calendarchik)

            time = QtWidgets.QDateTimeEdit()
            time.setDisplayFormat('HH:mm')
            time.setTime(QTime(12, 0))
            self.table_widget9999.setCellWidget(row_position - 1, 7, time)

            self.table_widget9999.setWordWrap(True)
            self.table_widget9999.resizeColumnsToContents()
            self.table_widget9999.setColumnWidth(0, 180)
            self.table_widget9999.setColumnWidth(1, 90)
            self.table_widget9999.setColumnWidth(2, 180)
            self.table_widget9999.setColumnWidth(5, 180)
            self.table_widget9999.setColumnWidth(6, 90)
            self.table_widget9999.resizeRowsToContents()
            self.table_widget9999.setColumnWidth(7, 90)

            log_print('Добавлено поле для ввода приказа')
        except Exception as e:
            log_print(e, 'Ошибка при добавлении поля для ввода приказа')
            pass

    def new_chelikikomic(self):
        try:
            current_row_count = self.table_widget99.rowCount()
            self.table_widget99.setColumnCount(4)
            self.table_widget99.setRowCount(current_row_count + 1)
            combo = QComboBox()
            combo.addItems(
                ["", "Назначения", "Председатель конкурсной комиссии", "Заместитель председателя комиссии",
                 "Секретарь конкурсной комиссии", "Член конкурсной комиссии", "Независимый эксперт",
                 "Наблюдатель"])
            self.table_widget99.setCellWidget(current_row_count, 3, combo)
            self.table_widget99.setCellWidget(current_row_count, 3, combo)

            selected_row = self.table_widget33.currentRow()
            row_items = []

            for column in range(self.table_widget33.columnCount()):
                item = self.table_widget33.item(selected_row, column)
            row_position = self.table_widget99.rowCount()

            self.table_widget99.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")
            self.table_widget99.setHorizontalHeaderLabels(
                ["Подразделение", "Должность", "ФИО", "Статус участника конкурсной комиссии"])

            self.table_widget99.setWordWrap(True)
            self.table_widget99.resizeColumnsToContents()
            self.table_widget99.setColumnWidth(0, 180)
            self.table_widget99.setColumnWidth(1, 90)
            self.table_widget99.setColumnWidth(2, 180)
            self.table_widget99.resizeRowsToContents()

            log_print('Добавлено поле для ввода приказа')
        except Exception as e:
            log_print(e, 'Ошибка при добавлении поля для ввода приказа')
            pass

    def search_tablerabot(self, text):
        currentTable = self.table_widget23
        # очистка выделения
        currentTable.clearSelection()
        # поиск и обновление видимости строк
        try:
            for i in range(currentTable.rowCount()):
                matches = False
                for j in range(currentTable.columnCount()):
                    item = currentTable.item(i, j)
                    if item is not None and text.lower() in item.text().lower():
                        matches = True
                        break
                currentTable.setRowHidden(i, not matches)
            log_print('Выполнен поиск в конкурсе')
        except Exception as e:
            log_print(f"Ошибка, при поиске в конкурсе: {e}")
            print(f"Ошибка, при поиске в конкурсе: {e}")

    def search_tablekomic(self, text):
        currentTable = self.table_widget33
        # очистка выделения
        currentTable.clearSelection()
        # поиск и обновление видимости строк
        try:
            for i in range(currentTable.rowCount()):
                matches = False
                for j in range(currentTable.columnCount()):
                    item = currentTable.item(i, j)
                    if item is not None and text.lower() in item.text().lower():
                        matches = True
                        break
                currentTable.setRowHidden(i, not matches)
            log_print('Выполнен поиск в конкурсе')
        except Exception as e:
            log_print(f"Ошибка, при поиске в конкурсе: {e}")
            print(f"Ошибка, при поиске в конкурсе: {e}")

    def zapoln_chelikrabot(self):
        df2 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
        # df33 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', usecols='N')
        # df3 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A')
        df2_filtered = df2.dropna(subset=[df2.columns[1]])
        df_orders = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Приказы в работе', usecols='C')

        df2_filtered['Закрыто'] = [0 for i in range(len(df2_filtered))]
        df2_filtered.loc[df2_filtered['ФИО'].isin(df_orders['ФИО']), 'Закрыто'] = 1
        df2_filtered.loc[df2_filtered['Unnamed: 0'] == 1.000, 'Закрыто'] = 2

        df2_filtered = df2_filtered.iloc[:, 1:]

        # Установка количества строк и столбцов в QTableWidget
        self.table_widget23.setRowCount(df2_filtered.shape[0])
        self.table_widget23.setColumnCount(3)

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df2_filtered.shape[0]):
            for col in range(df2_filtered.shape[1]):
                xxx = str(df2_filtered.iat[row, col])
                if xxx == '0':
                    xxx = ''
                elif xxx == '1':
                    xxx = 'Заведен конкурс'
                elif xxx == '2':
                    xxx = 'Должность на рассмотрении'
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget23.setItem(row, col, item)

        self.table_widget23.resizeColumnsToContents()
        self.table_widget23.setHorizontalHeaderLabels(["Подразделение", "Должность", "ФИО"])
        self.table_widget23.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        self.table_widget9999.setWordWrap(True)
        self.table_widget9999.resizeColumnsToContents()
        self.table_widget9999.setColumnWidth(0, 180)
        self.table_widget9999.setColumnWidth(1, 90)
        self.table_widget9999.setColumnWidth(2, 180)
        self.table_widget9999.setColumnWidth(5, 90)
        self.table_widget9999.resizeRowsToContents()

        self.table_widget23.setColumnWidth(3, 250)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()
        log_print('Таблица конкурс успешно заполнена')

    def zapoln_chelikkomic(self):
        df2 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A:D')
        # df33 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Лист1', usecols='N')
        # df3 = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Штатка', usecols='A')
        df2_filtered = df2.dropna(subset=[df2.columns[1]])
        df_orders = pd.read_excel('addition/kadrifile.xlsx', sheet_name='Приказы в работе', usecols='C')

        df2_filtered['Закрыто'] = [0 for i in range(len(df2_filtered))]
        df2_filtered.loc[df2_filtered['ФИО'].isin(df_orders['ФИО']), 'Закрыто'] = 1
        df2_filtered.loc[df2_filtered['Unnamed: 0'] == 1.000, 'Закрыто'] = 2

        dFilter = df2_filtered[
            ((df2_filtered.Должность == 'Руководитель') | (df2_filtered.Должность == 'Заместитель руководителя')
             | (df2_filtered.Должность == 'Начальник отдела') | (
                         df2_filtered.Должность == 'Заместитель начальника отдела')) & (~df2_filtered['ФИО'].isnull())]

        df2_filtered = dFilter.iloc[:, 1:]

        # Установка количества строк и столбцов в QTableWidget
        self.table_widget33.setRowCount(dFilter.shape[0])
        self.table_widget33.setColumnCount(3)

        # Заполнение QTableWidget данными из DataFrame
        for row in range(df2_filtered.shape[0]):
            for col in range(df2_filtered.shape[1]):
                xxx = str(df2_filtered.iat[row, col])
                if xxx == '0':
                    xxx = ''
                elif xxx == '1':
                    xxx = 'Заведен конкурс'
                elif xxx == '2':
                    xxx = 'Должность на рассмотрении'
                if xxx == 'nan':
                    xxx = ''
                item = QTableWidgetItem(xxx)
                self.table_widget33.setItem(row, col, item)

        self.table_widget33.resizeColumnsToContents()
        self.table_widget33.setHorizontalHeaderLabels(["Подразделение", "Должность", "ФИО"])
        self.table_widget33.horizontalHeader().setStyleSheet("font-size: 15px; font-weight: bold;")

        self.table_widget99.setWordWrap(True)
        self.table_widget99.resizeColumnsToContents()
        self.table_widget99.setColumnWidth(0, 180)
        self.table_widget99.setColumnWidth(1, 90)
        self.table_widget99.setColumnWidth(2, 180)
        self.table_widget99.setColumnWidth(5, 90)
        self.table_widget99.resizeRowsToContents()

        self.table_widget33.setColumnWidth(3, 250)
        excel = win32.Dispatch("Excel.Application")
        excel.Quit()
        log_print('Таблица конкурс успешно заполнена')

    def onExcelButton_click(self):
        print(123)